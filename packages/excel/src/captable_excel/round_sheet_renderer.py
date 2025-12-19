"""Round-style renderer (image-inspired) without Excel tables."""

from __future__ import annotations

from dataclasses import dataclass
from datetime import date
from decimal import Decimal
from typing import Dict, List, Optional, Tuple

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, Border, Side, Alignment, PatternFill
from openpyxl.comments import Comment
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.workbook.defined_name import DefinedName

from captable_domain.schemas import CapTableSnapshotCFG, WorkbookCFG, CapTableSnapshot


@dataclass
class HolderLine:
    holder_id: str
    shares: Decimal
    share_class_id: str
    investment: Optional[Decimal]


class RoundSheetRenderer:
    """Render one sheet per snapshot with columns per round (common, each preferred)."""

    def __init__(self, config: WorkbookCFG):
        self.config = config

        # Define styles
        self.blue_font = Font(color="0000FF")  # Blue for input values
        self.black_font = Font(color="000000")  # Black for calculated values
        self.green_font = Font(color="006400")  # Dark green for cross-sheet links
        self.bold_font = Font(bold=True)
        self.bold_blue_font = Font(bold=True, color="0000FF")

        # Header styling
        self.header_font = Font(bold=True, color="FFFFFF")  # White text on dark blue
        self.header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")  # Dark blue

        # Section header styling
        self.section_header_font = Font(italic=True, bold=True)
        self.section_header_fill = PatternFill(start_color="E7E6E6", end_color="E7E6E6", fill_type="solid")  # Light gray

        # Valuation row styling
        self.valuation_fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")  # Very light gray

        # Calculator styling
        self.calculator_fill = PatternFill(start_color="FFFACD", end_color="FFFACD", fill_type="solid")  # Light yellow
        self.calculator_header_font = Font(bold=True, size=11)
        self.calculator_label_font = Font(italic=True, size=10)

        # Interactive input cell styling (Phase 1: Essential)
        self.input_cell_fill = PatternFill(start_color="CCE5FF", end_color="CCE5FF", fill_type="solid")  # Light blue background
        self.input_cell_font = Font(bold=True, color="000000")  # Bold black text
        self.input_cell_border = Border(
            left=Side(style='medium', color='0070C0'),
            right=Side(style='medium', color='0070C0'),
            top=Side(style='medium', color='0070C0'),
            bottom=Side(style='medium', color='0070C0')
        )  # Medium blue border

        # Border styles
        self.thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        self.thick_border = Border(
            left=Side(style='medium'),
            right=Side(style='medium'),
            top=Side(style='medium'),
            bottom=Side(style='medium')
        )
        self.top_border = Border(top=Side(style='medium'))
        self.bottom_border = Border(bottom=Side(style='medium'))

        # Alignment
        self.center_align = Alignment(horizontal='center', vertical='center')
        self.right_align = Alignment(horizontal='right')

        # Track valuation cell positions per sheet for cross-sheet references
        # Key: (sheet_label, pref_id) -> {'pps': cell_ref, 'pre_money': cell_ref, 'post_money': cell_ref}
        self._valuation_cells: Dict[tuple, Dict[str, str]] = {}

        # Store column mappings per sheet for cross-sheet references
        # Key: sheet_label -> col_map dict
        self._sheet_col_maps: Dict[str, Dict[str, str]] = {}

    def render(self, output_path: str) -> str:
        wb = self.build_workbook()
        wb.save(output_path)
        return output_path

    def build_workbook(self) -> Workbook:
        wb = Workbook()
        wb.remove(wb.active)

        # Enable iterative calculation for round calculator formulas
        wb.calculation.calcMode = 'auto'
        wb.calculation.iterate = True
        wb.calculation.iterateCount = 100
        wb.calculation.iterateDelta = 0.001

        for idx, snap_cfg in enumerate(self.config.cap_table_snapshots):
            prev_label = self.config.cap_table_snapshots[idx - 1].label if idx > 0 else None
            prev_as_of_date = self.config.cap_table_snapshots[idx - 1].as_of_date if idx > 0 else None
            self._render_snapshot_sheet(wb, snap_cfg, prev_label, prev_as_of_date)

        return wb

    # ------------------------------------------------------------------ #
    # Allocation Mode Helper Methods
    # ------------------------------------------------------------------ #

    def _get_investor_allocation_mode(self, holder_id: str, calc_cfg) -> str:
        """Get the allocation mode for a specific investor.

        Args:
            holder_id: The holder/investor name
            calc_cfg: RoundCalculatorCFG instance

        Returns:
            Allocation mode: "manual", "target_ownership", or "pro_rata"
        """
        if not calc_cfg.enabled:
            return "manual"

        # Check for per-investor override
        if calc_cfg.per_investor_allocation and holder_id in calc_cfg.per_investor_allocation:
            return calc_cfg.per_investor_allocation[holder_id]

        # Use default allocation mode
        return calc_cfg.investment_allocation_mode

    def _get_investor_target_pct(self, holder_id: str, calc_cfg) -> Optional[float]:
        """Get the target ownership % for a specific investor.

        Args:
            holder_id: The holder/investor name
            calc_cfg: RoundCalculatorCFG instance

        Returns:
            Target ownership % (e.g., 0.20 for 20%) or None
        """
        # Check for per-investor target
        if calc_cfg.per_investor_target_pct and holder_id in calc_cfg.per_investor_target_pct:
            return calc_cfg.per_investor_target_pct[holder_id]

        # Fall back to default target
        return calc_cfg.target_ownership_pct

    def _generate_target_ownership_formula(
        self,
        target_pct: float,
        pre_money_cell_ref: str
    ) -> str:
        """Generate formula to calculate investment needed for target ownership %.

        Formula: Investment = (Target% × Pre-Money) / (1 - Target%)

        This avoids circular dependency since:
        - Post-Money = Pre-Money + Total Investment
        - Target% = Investment / Post-Money

        Solving algebraically:
        - Investment = (Target% × Pre-Money) / (1 - Target%)

        Args:
            target_pct: Target ownership % (e.g., 0.20 for 20%)
            pre_money_cell_ref: Excel cell reference for pre-money valuation

        Returns:
            Excel formula string
        """
        return f"=IFERROR(({target_pct}*{pre_money_cell_ref})/(1-{target_pct}),\"\")"

    def _generate_prorata_formula(
        self,
        holder_id: str,
        prev_total_shares_col: str,
        prev_totals_row: int,
        current_pps_cell_ref: str,
        current_total_shares_cell_ref: str
    ) -> str:
        """Generate formula to calculate investment to maintain previous ownership %.

        Formula: Investment = (Previous% * Current Total Shares) * PPS
        Where Previous% = Previous Holder Shares / Previous Total Shares

        Args:
            holder_id: The holder/investor name
            prev_total_shares_col: Excel column for previous total shares
            prev_totals_row: Row number for totals in previous round
            current_pps_cell_ref: Excel cell reference for current round PPS
            current_total_shares_cell_ref: Excel cell reference for current total shares

        Returns:
            Excel formula string
        """
        # This is a placeholder - actual implementation would need to track
        # the investor's row in previous round to calculate their previous %
        # For now, return a simplified formula
        return f"=IFERROR(0,\"\")"  # TODO: Implement pro-rata calculation

    # ------------------------------------------------------------------ #
    def _render_snapshot_sheet(self, wb: Workbook, snap_cfg: CapTableSnapshotCFG, prev_label: Optional[str] = None, prev_as_of_date: Optional[date] = None) -> None:
        snapshot = (
            snap_cfg.cap_table.snapshot(snap_cfg.as_of_date)
            if snap_cfg.as_of_date
            else snap_cfg.cap_table.current_snapshot()
        )

        # Excel sheet names must be <= 31 characters
        sheet_title = snap_cfg.label[:31] if len(snap_cfg.label) > 31 else snap_cfg.label
        sheet = wb.create_sheet(title=sheet_title)
        sheet.sheet_properties.pageSetUpPr.fitToPage = True
        sheet.sheet_view.showGridLines = False  # Hide gridlines

        title_cell = sheet["A1"]
        title_cell.value = f"Cap Table - {snap_cfg.label}"
        title_cell.font = Font(size=14, bold=True)
        title_cell.alignment = self.center_align

        # Identify classes present in positions (so Seed snapshot doesn't show Series A columns)
        common_class_ids = sorted(
            {p.share_class_id for p in snapshot.positions if not p.is_option and snapshot.share_classes[p.share_class_id].share_type == "common"}
        )
        pref_class_ids = sorted(
            {p.share_class_id for p in snapshot.positions if not p.is_option and snapshot.share_classes[p.share_class_id].share_type == "preferred"}
        )

        # Determine which rounds are from previous snapshot (should be shown in green)
        prev_pref_class_ids = set()
        if prev_label:
            # Find previous snapshot config
            for cfg in self.config.cap_table_snapshots:
                if cfg.label == prev_label:
                    prev_snapshot = cfg.cap_table.snapshot(cfg.as_of_date) if cfg.as_of_date else cfg.cap_table.current_snapshot()
                    prev_pref_class_ids = {p.share_class_id for p in prev_snapshot.positions if not p.is_option and prev_snapshot.share_classes[p.share_class_id].share_type == "preferred"}
                    break

        # Get option pool creation events per round
        # Option pools can be standalone events or nested in RoundClosingEvent.option_pool_created
        option_pool_by_round: Dict[str, Decimal] = {}
        for event in snap_cfg.cap_table.events:
            # Check for RoundClosingEvent which may contain option_pool_created
            if hasattr(event, 'option_pool_created') and event.option_pool_created:
                pool = event.option_pool_created
                # Associate with the round's share class (look for share_issuances)
                if hasattr(event, 'share_issuances') and event.share_issuances:
                    for issuance in event.share_issuances:
                        if issuance.share_class_id in pref_class_ids:
                            option_pool_by_round[issuance.share_class_id] = option_pool_by_round.get(issuance.share_class_id, Decimal("0")) + pool.shares_authorized
                            break
            # Also check for standalone OptionPoolCreation events
            elif hasattr(event, 'shares_authorized') and hasattr(event, 'event_date'):
                # Find which round this option pool expansion belongs to
                event_date = event.event_date
                # Associate with the preferred round that happened around the same time
                for pref_id in pref_class_ids:
                    # Find events for this preferred class
                    pref_events = [e for e in snap_cfg.cap_table.events
                                 if hasattr(e, 'share_class_id') and e.share_class_id == pref_id]
                    if pref_events:
                        # If option pool event is within 60 days of first pref event, associate it
                        first_pref_date = min(e.event_date for e in pref_events)
                        if abs((event_date - first_pref_date).days) <= 60:
                            option_pool_by_round[pref_id] = option_pool_by_round.get(pref_id, Decimal("0")) + event.shares_authorized

        # Identify SAFE conversion rounds (they have conversion_price, not pre-money)
        # SAFEConversionEvents are nested inside RoundClosingEvent.safe_conversions
        # Store valuation_cap and discount_rate to build formulas
        # Only include SAFE rounds that:
        # 1. Occurred on or before this snapshot's date
        # 2. Are NEW in this snapshot (not already shown in previous snapshot)
        safe_rounds: Dict[str, dict] = {}  # share_class_id -> {valuation_cap, discount_rate}
        for event in snap_cfg.cap_table.events:
            # Filter by snapshot date
            if snap_cfg.as_of_date and event.event_date > snap_cfg.as_of_date:
                continue
            # Check for RoundClosingEvent which contains safe_conversions
            if hasattr(event, 'safe_conversions') and event.safe_conversions:
                for safe_conv in event.safe_conversions:
                    # Each SAFEConversionEvent has resulting_share_class_id and safe_instrument
                    share_class_id = safe_conv.resulting_share_class_id
                    # Only include if this is a NEW round (not in previous snapshot)
                    if share_class_id not in prev_pref_class_ids:
                        safe_rounds[share_class_id] = {
                            'valuation_cap': safe_conv.safe_instrument.valuation_cap,
                            'discount_rate': safe_conv.safe_instrument.discount_rate,
                        }
            # Also check for standalone SAFEConversionEvent (just in case)
            elif hasattr(event, 'event_type') and event.event_type == 'safe_conversion':
                if snap_cfg.as_of_date and event.event_date > snap_cfg.as_of_date:
                    continue
                share_class_id = event.resulting_share_class_id
                if share_class_id not in prev_pref_class_ids:
                    safe_rounds[share_class_id] = {
                        'valuation_cap': event.safe_instrument.valuation_cap,
                        'discount_rate': event.safe_instrument.discount_rate,
                    }

        # Identify secondary transactions (ShareTransferEvents)
        # Only include transactions that are NEW in this snapshot:
        # 1. Occurred on or before this snapshot's date
        # 2. Occurred AFTER the previous snapshot's date (so they're new)
        # Group by: from_holder, to_holder, share_class -> shares, price
        # Support alchemy: resulting_share_class_id (buyer may get different class)
        secondary_transactions: List[dict] = []
        for event in snap_cfg.cap_table.events:
            if hasattr(event, 'event_type') and event.event_type == 'share_transfer':
                # Filter by snapshot date - only include if event is before/on as_of_date
                if snap_cfg.as_of_date and event.event_date > snap_cfg.as_of_date:
                    continue  # Skip events after the snapshot date
                # Filter to only NEW transactions (after previous snapshot date)
                if prev_as_of_date and event.event_date <= prev_as_of_date:
                    continue  # Skip events that were already in previous snapshot
                # Get resulting class (alchemy) - buyer may get different class than seller gave
                resulting_class = getattr(event, 'resulting_share_class_id', None) or event.share_class_id
                secondary_transactions.append({
                    'from_holder': event.from_holder_id,
                    'to_holder': event.to_holder_id,
                    'share_class': event.share_class_id,  # Seller's class
                    'resulting_class': resulting_class,   # Buyer's class (may differ)
                    'shares': event.shares,
                    'price_per_share': event.price_per_share,
                    'event_date': event.event_date,
                })


        # Build holder lines
        common_lines: List[HolderLine] = []
        pref_lines: List[HolderLine] = []
        for pos in snapshot.positions:
            if pos.is_option:
                continue
            if pos.share_class_id in common_class_ids:
                common_lines.append(HolderLine(pos.holder_id, pos.shares, pos.share_class_id, pos.cost_basis))
            elif pos.share_class_id in pref_class_ids:
                pref_lines.append(HolderLine(pos.holder_id, pos.shares, pos.share_class_id, pos.cost_basis))

        # Header columns: Common (# shares), GAP, then each preferred class ($ invested, Preferred shares, Option Pool), GAP, Total Shares/%
        col_map: Dict[str, str] = {}
        col_idx = 2  # start column B

        # Set header row height
        sheet.row_dimensions[3].height = 30

        # Column A header
        holder_header = sheet.cell(row=3, column=1, value="Holder")
        holder_header.font = self.header_font
        holder_header.fill = self.header_fill
        holder_header.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        holder_header.border = Border(right=Side(style='thin', color="FFFFFF"))

        # Common shares column
        header_cell = sheet.cell(row=3, column=col_idx, value="Common\nShares")
        header_cell.font = self.header_font
        header_cell.fill = self.header_fill
        header_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        header_cell.border = Border(right=Side(style='thin', color="FFFFFF"))
        col_map["common_shares"] = self._col_letter(col_idx)
        col_idx += 1

        # GAP after common
        gap_col_letter = self._col_letter(col_idx)
        sheet.column_dimensions[gap_col_letter].width = 3  # Narrow gap
        col_idx += 1

        # Preferred rounds (each with $Invested, Preferred Shares, and optionally Option Pool)
        # Track which rounds have option pools
        rounds_with_pools = set(option_pool_by_round.keys())

        for pref_id in pref_class_ids:
            inv_header = sheet.cell(row=3, column=col_idx, value=f"{pref_id}\n$ Invested")
            inv_header.font = self.header_font
            inv_header.fill = self.header_fill
            inv_header.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            inv_header.border = Border(right=Side(style='thin', color="FFFFFF"))
            col_map[f"{pref_id}_invested"] = self._col_letter(col_idx)
            col_idx += 1

            sh_header = sheet.cell(row=3, column=col_idx, value=f"{pref_id}\nPreferred")
            sh_header.font = self.header_font
            sh_header.fill = self.header_fill
            sh_header.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            sh_header.border = Border(right=Side(style='thin', color="FFFFFF"))
            col_map[f"{pref_id}_shares"] = self._col_letter(col_idx)
            col_idx += 1

            # Option Pool column only for rounds that have pools
            if pref_id in rounds_with_pools:
                opt_header = sheet.cell(row=3, column=col_idx, value=f"{pref_id}\nOption Pool")
                opt_header.font = self.header_font
                opt_header.fill = self.header_fill
                opt_header.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                opt_header.border = Border(right=Side(style='thin', color="FFFFFF"))
                col_map[f"{pref_id}_option_pool"] = self._col_letter(col_idx)
                col_idx += 1

            # GAP after each preferred round
            gap_col_letter = self._col_letter(col_idx)
            sheet.column_dimensions[gap_col_letter].width = 3  # Narrow gap
            col_idx += 1

        # GAP before summary columns (already added after each round)

        # Total Shares (FD) - consolidated column
        tot_sh_header = sheet.cell(row=3, column=col_idx, value="Total Shares\n(Fully Diluted)")
        tot_sh_header.font = self.header_font
        tot_sh_header.fill = self.header_fill
        tot_sh_header.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        tot_sh_header.border = Border(right=Side(style='thin', color="FFFFFF"))
        col_map["total_shares"] = self._col_letter(col_idx)
        col_idx += 1

        # % FD
        pct_fd_header = sheet.cell(row=3, column=col_idx, value="% Ownership\n(FD)")
        pct_fd_header.font = self.header_font
        pct_fd_header.fill = self.header_fill
        pct_fd_header.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        col_map["pct_fd"] = self._col_letter(col_idx)

        # Freeze panes at row 4 (after headers) and column B (after holder names)
        sheet.freeze_panes = "B4"

        # Common holder lines
        row = 4
        for line in common_lines:
            holder_cell = sheet[f"A{row}"]
            holder_cell.value = line.holder_id

            shares_cell = sheet[f"{col_map['common_shares']}{row}"]
            if prev_label:
                # Value with comment indicating it's from previous round
                # Use actual value but color green to show it's linked conceptually
                shares_cell.value = float(line.shares)
                shares_cell.font = self.green_font  # Indicates value from previous round
                shares_cell.comment = Comment(f"Value from {prev_label}", "System")
            else:
                # Hardcoded value for first snapshot
                shares_cell.value = float(line.shares)
                shares_cell.font = self.blue_font  # Hardcoded value
            shares_cell.number_format = '#,##0'
            row += 1

        # EMPTY ROW after common holders
        row += 1

        # Option Pool section (2 rows: Allocated Options and ESOP Available)
        allocated_row = row
        available_row = row + 1

        # Allocated Options row
        alloc_label = sheet[f"A{allocated_row}"]
        alloc_label.value = "Allocated Options"
        alloc_label.font = self.section_header_font
        alloc_label.fill = self.section_header_fill
        alloc_label.comment = Comment("Options exercised/granted (now held as shares)", "System")

        # Note: Allocated Options are tracked per-round in their respective option pool columns
        # Common Shares column is left empty for this row

        # ESOP Available row
        avail_label = sheet[f"A{available_row}"]
        avail_label.value = "ESOP Available"
        avail_label.font = self.section_header_font
        avail_label.fill = self.section_header_fill
        avail_label.comment = Comment("Options available for future grants", "System")

        # Note: ESOP Available is tracked per-round in their respective option pool columns
        # Common Shares column is left empty for this row

        # Option pool values are populated by the Option Pool Editor box
        # The main table cells will reference the editor input cells

        row = available_row + 1

        # EMPTY ROW after option pool
        row += 1

        # Preferred section header - add separator border
        pref_header_row = row
        pref_header_cell = sheet[f"A{pref_header_row}"]
        pref_header_cell.value = "Preferred Rounds"
        pref_header_cell.font = self.section_header_font
        pref_header_cell.fill = self.section_header_fill
        pref_header_cell.border = Border(top=Side(style='medium'))
        row += 1

        # Preferred holder lines
        pref_start_row = row

        # Determine which round the calculator applies to
        calc_cfg = snap_cfg.round_calculator
        target_round_id = calc_cfg.target_round_id if calc_cfg.enabled else None
        if target_round_id is None and calc_cfg.enabled and pref_class_ids:
            # Default to most recent preferred round
            target_round_id = pref_class_ids[-1]

        # Track cells that need formula updates (can't set attributes on Cell objects)
        cells_needing_target_formula = {}  # {(col, row): holder_id}
        cells_needing_prorata_formula = {}  # {(col, row): holder_id}

        # Build a mapping of holder_id -> {share_class_id: HolderLine}
        # This ensures each holder appears on ONE row with all their positions
        holder_positions: Dict[str, Dict[str, HolderLine]] = {}
        for line in pref_lines:
            if line.holder_id not in holder_positions:
                holder_positions[line.holder_id] = {}
            holder_positions[line.holder_id][line.share_class_id] = line

        # Identify which (holder, class) combinations are ONLY from secondary (no primary investment)
        # These should not have Investment/Shares calculated from primary formula
        # With alchemy, buyer gets resulting_class, not seller's share_class
        secondary_only_positions: set = set()
        for txn in secondary_transactions:
            # Buyer receives resulting_class (may differ from seller's share_class with alchemy)
            buyer_class = txn['resulting_class']
            buyer_key = (txn['to_holder'], buyer_class)
            # Check if buyer has any primary investment in this class
            # (i.e., they appear in share issuances or SAFE conversions for this class)
            has_primary = False
            for event in snap_cfg.cap_table.events:
                if hasattr(event, 'holder_id') and event.holder_id == txn['to_holder']:
                    if hasattr(event, 'share_class_id') and event.share_class_id == buyer_class:
                        has_primary = True
                        break
                # Check SAFE conversions in RoundClosingEvent
                if hasattr(event, 'safe_conversions'):
                    for safe_conv in event.safe_conversions:
                        if (safe_conv.safe_holder_id == txn['to_holder'] and
                            safe_conv.resulting_share_class_id == buyer_class):
                            has_primary = True
                            break
                # Check share issuances in RoundClosingEvent
                if hasattr(event, 'share_issuances'):
                    for issuance in event.share_issuances:
                        if (issuance.holder_id == txn['to_holder'] and
                            issuance.share_class_id == buyer_class):
                            has_primary = True
                            break
            if not has_primary:
                secondary_only_positions.add(buyer_key)

        # Get unique holders in order of first appearance
        unique_holders = list(holder_positions.keys())

        # Create one row per unique holder
        for holder_id in unique_holders:
            holder_cell = sheet[f"A{row}"]
            holder_cell.value = holder_id

            # Fill in each preferred class column for this holder
            for pref_id in pref_class_ids:
                invest_col = col_map[f"{pref_id}_invested"]
                shares_col = col_map[f"{pref_id}_shares"]

                # Get the position for this holder in this class (if any)
                line = holder_positions[holder_id].get(pref_id)

                # Determine if this round is from previous snapshot
                is_prev_round = pref_id in prev_pref_class_ids

                # Check if calculator applies to this round
                is_calc_round = (pref_id == target_round_id)

                # Investment cell
                invest_cell = sheet[f"{invest_col}{row}"]

                # Check if this is a secondary-only position (no primary investment)
                is_secondary_only = (holder_id, pref_id) in secondary_only_positions

                if line and not is_secondary_only:
                    # Holder has a position in this class
                    if is_prev_round:
                        # Previous round - always hardcoded (green)
                        invest_cell.value = float(line.investment) if line.investment is not None else None
                        invest_cell.font = self.green_font
                        invest_cell.comment = Comment(f"Value from {prev_label}", "System")
                    elif is_calc_round:
                        # Calculator round - blue interactive input
                        invest_cell.value = float(line.investment) if line.investment is not None else None
                        invest_cell.number_format = '$#,##0'
                        self._mark_as_input_cell(
                            invest_cell,
                            f"Investment amount for {holder_id} in {pref_id}.\n\n"
                            "Edit this value to change the investment.\n"
                            "Share count will update automatically based on price per share."
                        )
                    else:
                        # Non-calculator round - blue interactive input
                        invest_cell.value = float(line.investment) if line.investment is not None else None
                        invest_cell.number_format = '$#,##0'
                        self._mark_as_input_cell(
                            invest_cell,
                            f"Investment amount for {holder_id} in {pref_id}.\n\n"
                            "Edit this value to change the investment.\n"
                            "Share count will update automatically based on price per share."
                        )
                # else: no position in this class, leave cell empty

                # Shares cell (formula applied after PPS rows are defined)
                shares_cell = sheet[f"{shares_col}{row}"]
                shares_cell.value = None
                shares_cell.number_format = '#,##0'

            row += 1

        # EMPTY ROW after preferred holders
        row += 1

        totals_row = row
        totals_label = sheet[f"A{totals_row}"]
        totals_label.value = "Totals"
        totals_label.font = self.bold_font
        totals_label.border = Border(top=Side(style='medium'), bottom=Side(style='medium'))

        # Totals formulas
        common_sum_range = f"{col_map['common_shares']}4:{col_map['common_shares']}{pref_header_row - 1}"
        common_total = sheet[f"{col_map['common_shares']}{totals_row}"]
        common_total.value = f"=SUM({common_sum_range})"
        common_total.font = self.black_font  # Calculated
        common_total.border = Border(top=Side(style='medium'), bottom=Side(style='medium'))
        common_total.number_format = '#,##0'

        for pref_id in pref_class_ids:
            shares_col = col_map[f"{pref_id}_shares"]
            invest_col = col_map[f"{pref_id}_invested"]
            sum_range_sh = f"{shares_col}{pref_start_row}:{shares_col}{row-1}"
            sum_range_inv = f"{invest_col}{pref_start_row}:{invest_col}{row-1}"

            shares_total = sheet[f"{shares_col}{totals_row}"]
            shares_total.value = f"=SUM({sum_range_sh})"
            shares_total.font = self.black_font  # Calculated
            shares_total.border = Border(top=Side(style='medium'), bottom=Side(style='medium'))
            shares_total.number_format = '#,##0'

            invest_total = sheet[f"{invest_col}{totals_row}"]
            invest_total.value = f"=SUM({sum_range_inv})"
            invest_total.font = self.black_font  # Calculated
            invest_total.border = Border(top=Side(style='medium'), bottom=Side(style='medium'))
            invest_total.number_format = '$#,##0'

            # Option pool total for this round (only if round has pool)
            if pref_id in rounds_with_pools:
                opt_col = col_map[f"{pref_id}_option_pool"]
                opt_total = sheet[f"{opt_col}{totals_row}"]
                opt_total.value = f"=IFERROR({opt_col}{allocated_row}+{opt_col}{available_row},0)"
                opt_total.font = self.black_font  # Calculated
                opt_total.border = Border(top=Side(style='medium'), bottom=Side(style='medium'))
                opt_total.number_format = '#,##0'

        # Total Shares (FD) = sum of common + preferred + option pools (only for rounds with pools)
        components = [f"{col_map['common_shares']}{totals_row}"] + [
            f"{col_map[f'{pid}_shares']}{totals_row}" for pid in pref_class_ids
        ] + [f"{col_map[f'{pid}_option_pool']}{totals_row}" for pid in pref_class_ids if pid in rounds_with_pools]

        total_shares_totals = sheet[f"{col_map['total_shares']}{totals_row}"]
        total_shares_totals.value = f"={'+'.join(components)}"
        total_shares_totals.font = self.black_font  # Calculated
        total_shares_totals.border = Border(top=Side(style='medium'), bottom=Side(style='medium'))
        total_shares_totals.number_format = '#,##0'

        # Add % FD total (100%) in totals row
        pct_fd_total = sheet[f"{col_map['pct_fd']}{totals_row}"]
        pct_fd_total.value = 1.0
        pct_fd_total.font = self.black_font  # Should equal 100%
        pct_fd_total.border = Border(top=Side(style='medium'), bottom=Side(style='medium'))
        pct_fd_total.number_format = '0.0%'

        # Price per share / pre-money / post-money rows (per class)
        # Add separate row for starting shares to avoid confusion
        starting_shares_row = totals_row + 1
        pre_row = totals_row + 2
        price_row = totals_row + 3
        post_row = totals_row + 4

        start_label = sheet[f"A{starting_shares_row}"]
        start_label.value = "Starting Shares (pre-round)"
        start_label.font = Font(italic=True)
        start_label.fill = self.valuation_fill
        start_label.border = Border(top=Side(style='thin'))

        pre_label = sheet[f"A{pre_row}"]
        pre_label.value = "Pre-Money Valuation"
        pre_label.font = self.bold_font
        pre_label.fill = self.valuation_fill
        pre_label.border = Border(left=Side(style='thin'), top=Side(style='thin'))

        price_label = sheet[f"A{price_row}"]
        price_label.value = "Price per Share"
        price_label.font = self.bold_font
        price_label.fill = self.valuation_fill
        price_label.border = Border(left=Side(style='thin'))

        post_label = sheet[f"A{post_row}"]
        post_label.value = "Post-Money Valuation"
        post_label.font = self.bold_font
        post_label.fill = self.valuation_fill
        post_label.border = Border(left=Side(style='thin'), bottom=Side(style='thin'))

        pps_cells: Dict[str, str] = {}
        # Build starting shares per class (cumulative before that class)
        start_shares_cells: Dict[str, str] = {}
        common_total_cell = f"{col_map['common_shares']}{totals_row}"
        for idx, pref_id in enumerate(pref_class_ids):
            prev_pref_totals = [
                f"{col_map[f'{pid}_shares']}{totals_row}" for pid in pref_class_ids[:idx]
            ]
            prev_option_totals = [
                f"{col_map[f'{pid}_option_pool']}{totals_row}" for pid in pref_class_ids[:idx]
                if pid in rounds_with_pools
            ]
            start_formula_parts = [common_total_cell] + prev_pref_totals + prev_option_totals
            # Put starting shares in the STARTING SHARES row, in the shares column
            start_cell_ref = f"{col_map[f'{pref_id}_shares']}{starting_shares_row}"
            start_cell = sheet[start_cell_ref]
            start_cell.value = f"={'+'.join(start_formula_parts)}"
            start_cell.font = self.black_font  # Calculated
            start_cell.number_format = '#,##0'
            start_shares_cells[pref_id] = start_cell_ref

        # Calculate pre-money values from actual data (reverse-engineer from investments and shares)
        pref_pre_money: Dict[str, Optional[Decimal]] = {}
        for pref_id in pref_class_ids:
            # Get total investment and shares for this class
            class_lines = [line for line in pref_lines if line.share_class_id == pref_id]
            if not class_lines:
                pref_pre_money[pref_id] = None
                continue

            total_investment = sum(line.investment for line in class_lines if line.investment)
            total_shares = sum(line.shares for line in class_lines)

            if total_investment and total_shares and total_investment > 0 and total_shares > 0:
                # PPS = total investment / total shares
                pps = total_investment / total_shares

                # Pre-money = PPS * (shares before this round)
                # Calculate shares before this round
                idx = pref_class_ids.index(pref_id)
                shares_before = (
                    sum(line.shares for line in common_lines) +
                    sum(option_pool_by_round.get(pid, Decimal("0")) for pid in pref_class_ids[:idx]) +
                    sum(sum(l.shares for l in pref_lines if l.share_class_id == pid)
                        for pid in pref_class_ids[:idx])
                )

                pre_money = pps * shares_before
                pref_pre_money[pref_id] = pre_money
            else:
                pref_pre_money[pref_id] = None

        for idx, pref_id in enumerate(pref_class_ids):
            invest_col = col_map[f"{pref_id}_invested"]
            shares_col = col_map[f"{pref_id}_shares"]
            start_cell = start_shares_cells[pref_id]

            pre_money_cell_ref = f"{invest_col}{pre_row}"
            pps_cell_ref = f"{invest_col}{price_row}"
            post_money_cell_ref = f"{invest_col}{post_row}"

            pps_cells[pref_id] = pps_cell_ref

            # Determine if this is the first or last preferred round for border styling
            is_first_pref = (idx == 0)
            is_last_pref = (idx == len(pref_class_ids) - 1)

            # Determine if this round is from previous snapshot
            is_prev_round = pref_id in prev_pref_class_ids

            # Determine if this is a SAFE conversion (no pre/post money)
            is_safe_round = pref_id in safe_rounds

            # Starting shares cell
            start_cell_obj = sheet[start_cell]
            start_cell_obj.fill = self.valuation_fill
            start_cell_obj.border = Border(
                left=Side(style='thin') if is_first_pref else None,
                right=Side(style='thin') if is_last_pref else None,
                top=Side(style='thin')
            )

            # Look up previous sheet's cell refs for this round (if referencing previous sheet)
            prev_cells = self._valuation_cells.get((prev_label, pref_id), {}) if prev_label else {}

            if is_safe_round:
                # SAFE round: PPS formula will reference SAFE editor box
                # Pre-money cell is empty for SAFEs
                pre_money_cell = sheet[pre_money_cell_ref]
                pre_money_cell.value = None
                pre_money_cell.fill = self.valuation_fill
                pre_money_cell.border = Border(
                    left=Side(style='thin') if is_first_pref else None,
                    right=Side(style='thin') if is_last_pref else None,
                    top=Side(style='thin')
                )

                # PPS: reference previous sheet if this is a prior round, otherwise SAFE editor sets it
                pps_cell = sheet[pps_cell_ref]
                if is_prev_round and prev_label and 'pps' in prev_cells:
                    # Reference previous sheet's actual PPS cell (may differ in row number)
                    prev_pps_ref = prev_cells['pps']
                    pps_cell.value = f"='{prev_label}'!{prev_pps_ref}"
                    pps_cell.font = self.green_font
                    pps_cell.comment = Comment(f"PPS from {prev_label} (preserved despite alchemy)", "System")
                else:
                    pps_cell.value = None  # Will be formula referencing SAFE editor
                    pps_cell.font = self.black_font
                pps_cell.number_format = '$0.00'
                pps_cell.fill = self.valuation_fill
                pps_cell.border = Border(
                    left=Side(style='thin') if is_first_pref else None,
                    right=Side(style='thin') if is_last_pref else None
                )

                # Post-money: SAFEs don't have post-money (leave empty)
                post_money_cell = sheet[post_money_cell_ref]
                post_money_cell.value = None
                post_money_cell.fill = self.valuation_fill
                post_money_cell.border = Border(
                    left=Side(style='thin') if is_first_pref else None,
                    right=Side(style='thin') if is_last_pref else None,
                    bottom=Side(style='thin')
                )
            else:
                # Priced round: Pre-money, PPS, Post-money
                pre_money_cell = sheet[pre_money_cell_ref]
                pps_cell = sheet[pps_cell_ref]
                post_money_cell = sheet[post_money_cell_ref]

                if is_prev_round and prev_label and prev_cells:
                    # Reference previous sheet's actual cell positions (may differ in row numbers)
                    prev_pre_ref = prev_cells.get('pre_money', pre_money_cell_ref)
                    prev_pps_ref = prev_cells.get('pps', pps_cell_ref)
                    prev_post_ref = prev_cells.get('post_money', post_money_cell_ref)

                    pre_money_cell.value = f"='{prev_label}'!{prev_pre_ref}"
                    pre_money_cell.font = self.green_font
                    pre_money_cell.comment = Comment(f"Pre-money from {prev_label} (preserved despite alchemy)", "System")

                    pps_cell.value = f"='{prev_label}'!{prev_pps_ref}"
                    pps_cell.font = self.green_font
                    pps_cell.comment = Comment(f"PPS from {prev_label} (preserved despite alchemy)", "System")

                    post_money_cell.value = f"='{prev_label}'!{prev_post_ref}"
                    post_money_cell.font = self.green_font
                    post_money_cell.comment = Comment(f"Post-money from {prev_label} (preserved despite alchemy)", "System")
                else:
                    # Current round: editable pre-money, calculated PPS and post-money
                    if pref_pre_money.get(pref_id) is not None:
                        pre_money_cell.value = float(pref_pre_money[pref_id])
                        self._mark_as_input_cell(
                            pre_money_cell,
                            f"Pre-money valuation for {pref_id}.\n\n"
                            "Edit this value to change the pre-money valuation.\n"
                            "Price per share and ownership % will update automatically."
                        )
                    else:
                        pre_money_cell.value = None
                        self._mark_as_input_cell(
                            pre_money_cell,
                            f"Pre-money valuation for {pref_id}.\n\n"
                            "Edit this value to set the pre-money valuation.\n"
                            "Price per share and ownership % will update automatically."
                        )

                    # PPS = pre-money / starting shares
                    pps_cell.value = f"=IFERROR({pre_money_cell_ref}/{start_cell},\"\")"
                    pps_cell.font = self.black_font

                    # Post-money = pre-money + total invested for this class
                    post_money_cell.value = f"=IFERROR({pre_money_cell_ref}+{invest_col}{totals_row},\"\")"
                    post_money_cell.font = self.black_font

                pre_money_cell.number_format = '$#,##0'
                pre_money_cell.border = Border(
                    left=Side(style='thin') if is_first_pref else None,
                    right=Side(style='thin') if is_last_pref else None,
                    top=Side(style='thin')
                )

                pps_cell.number_format = '$0.00'
                pps_cell.fill = self.valuation_fill
                pps_cell.border = Border(
                    left=Side(style='thin') if is_first_pref else None,
                    right=Side(style='thin') if is_last_pref else None
                )

                post_money_cell.number_format = '$#,##0'
                post_money_cell.fill = self.valuation_fill
                post_money_cell.border = Border(
                    left=Side(style='thin') if is_first_pref else None,
                    right=Side(style='thin') if is_last_pref else None,
                    bottom=Side(style='thin')
                )

            # Store this sheet's valuation cell refs for future sheets to reference
            self._valuation_cells[(snap_cfg.label, pref_id)] = {
                'pps': pps_cell_ref,
                'pre_money': pre_money_cell_ref,
                'post_money': post_money_cell_ref,
            }

            # Apply per-holder shares = investment / PPS
            # Only add formula if the investor actually has an investment in this round
            for r in range(pref_start_row, row):
                invest_value = sheet[f"{invest_col}{r}"].value
                # Only add shares formula if there's an investment amount
                if invest_value is not None and invest_value != "":
                    shares_formula_cell = sheet[f"{shares_col}{r}"]
                    shares_formula_cell.value = f"=IFERROR({invest_col}{r}/{pps_cell_ref},0)"
                    if is_prev_round:
                        shares_formula_cell.font = self.green_font  # From previous round (still formula but green)
                    else:
                        shares_formula_cell.font = self.black_font  # Calculated from investment/PPS

            # Second pass: Update investment cells with formulas for non-manual allocation modes
            if is_calc_round and pref_id == target_round_id:
                # Process target ownership formulas
                for (col, r), holder_id in cells_needing_target_formula.items():
                    if col == invest_col:  # Only process cells for this round
                        invest_cell = sheet.cell(row=r, column=self._col_to_index(col))
                        target_pct = self._get_investor_target_pct(holder_id, calc_cfg)

                        if target_pct is not None:
                            # Generate formula: Investment = (Target% × Pre-Money) / (1 - Target%)
                            formula = self._generate_target_ownership_formula(target_pct, pre_money_cell_ref)
                            invest_cell.value = formula
                            invest_cell.font = self.black_font  # Formula-driven (black)

                # Process pro-rata formulas
                for (col, r), holder_id in cells_needing_prorata_formula.items():
                    if col == invest_col:  # Only process cells for this round
                        invest_cell = sheet.cell(row=r, column=self._col_to_index(col))

                        # Generate pro-rata formula
                        # TODO: Implement actual pro-rata calculation
                        # For now, keep as hardcoded value
                        invest_cell.font = self.blue_font

        # Total Shares and % FD per holder rows
        share_columns = [col_map["common_shares"]] + [col_map[f"{pid}_shares"] for pid in pref_class_ids] + [col_map[f"{pid}_option_pool"] for pid in pref_class_ids if pid in rounds_with_pools]

        for r in range(4, row):
            # Get the holder name in column A
            cell_a_value = sheet.cell(row=r, column=1).value

            # Skip empty rows, header rows, and section separators
            if not cell_a_value:
                continue
            if r == pref_header_row:
                continue
            if cell_a_value == "Preferred Rounds":
                continue

            # Handle option pool rows specially
            if cell_a_value in ["Allocated Options", "ESOP Available"]:
                # Still calculate Total Shares for option pool rows
                share_sums = "+".join(f"{col}{r}" for col in share_columns)
                total_shares_cell = sheet[f"{col_map['total_shares']}{r}"]
                total_shares_cell.value = f"=IFERROR({share_sums},\"\")"  # Blank if 0
                total_shares_cell.font = self.black_font  # Calculated
                total_shares_cell.number_format = '#,##0'

                # % FD for option pool rows
                pct_fd_cell = sheet[f"{col_map['pct_fd']}{r}"]
                pct_fd_cell.value = f"=IFERROR({col_map['total_shares']}{r}/{col_map['total_shares']}{totals_row},\"\")"  # Blank if 0
                pct_fd_cell.font = self.black_font  # Calculated
                pct_fd_cell.number_format = '0.0%'
                continue

            # Total Shares = sum of all share columns for this holder
            share_sums = "+".join(f"{col}{r}" for col in share_columns)
            total_shares_cell = sheet[f"{col_map['total_shares']}{r}"]
            total_shares_cell.value = f"=IFERROR({share_sums},\"\")"  # Blank if 0
            total_shares_cell.font = self.black_font  # Calculated
            total_shares_cell.number_format = '#,##0'

            # % FD = Total Shares / Total Shares in totals row
            pct_fd_cell = sheet[f"{col_map['pct_fd']}{r}"]
            pct_fd_cell.value = f"=IFERROR({col_map['total_shares']}{r}/{col_map['total_shares']}{totals_row},\"\")"  # Blank if 0
            pct_fd_cell.font = self.black_font  # Calculated
            pct_fd_cell.number_format = '0.0%'

        # Adjust column widths for better readability
        sheet.column_dimensions['A'].width = 25  # Holder names
        sheet.column_dimensions[col_map['common_shares']].width = 15  # Common shares
        for pref_id in pref_class_ids:
            sheet.column_dimensions[col_map[f"{pref_id}_invested"]].width = 15  # $ Invested
            sheet.column_dimensions[col_map[f"{pref_id}_shares"]].width = 15  # Preferred shares
            if pref_id in rounds_with_pools:
                sheet.column_dimensions[col_map[f"{pref_id}_option_pool"]].width = 15  # Option pool
        sheet.column_dimensions[col_map['total_shares']].width = 15  # Total shares (FD)
        sheet.column_dimensions[col_map['pct_fd']].width = 12  # % FD

        # Add Option Pool Editor box to the right of the table (only if there are rounds with pools)
        if rounds_with_pools:
            self._render_option_pool_editor(
                sheet, col_map, pref_class_ids, rounds_with_pools,
                allocated_row, available_row, prev_pref_class_ids, prev_label, totals_row
            )

        # Add SAFE Editor box below the cap table (only if there are SAFEs)
        safe_editor_end_row = post_row
        if safe_rounds:
            safe_editor_end_row = self._render_safe_editor(
                sheet, col_map, pref_class_ids, safe_rounds,
                post_row, start_shares_cells, pps_cells,
                prev_pref_class_ids, prev_label
            )

        # Add Pro Rata Editor box below SAFE editor
        # Show when there's a previous snapshot and a new priced round (not SAFE)
        pro_rata_end_row = safe_editor_end_row
        new_priced_rounds = [pid for pid in pref_class_ids if pid not in safe_rounds and pid not in prev_pref_class_ids]
        if prev_label and new_priced_rounds:
            # Get the newest priced round for pro rata
            pro_rata_round_id = new_priced_rounds[-1]

            # Get previous snapshot holder ownership percentages
            # Only include holders whose share class has pro rata rights
            # Use same basis as cap table % column (total_shares_outstanding, excludes pool)
            # This ensures the pro rata % matches what the investor sees as their ownership
            prev_snapshot_holders: List[dict] = []
            for cfg in self.config.cap_table_snapshots:
                if cfg.label == prev_label:
                    prev_snap = cfg.cap_table.snapshot(cfg.as_of_date) if cfg.as_of_date else cfg.cap_table.current_snapshot()
                    prev_total = prev_snap.total_shares_outstanding
                    for pos in prev_snap.positions:
                        if pos.is_option or prev_total <= 0:
                            continue
                        # Check if share class has pro rata rights
                        share_class = prev_snap.share_classes.get(pos.share_class_id)
                        has_pro_rata = share_class and getattr(share_class, 'has_pro_rata_rights', False)
                        if has_pro_rata:
                            prev_snapshot_holders.append({
                                'holder_id': pos.holder_id,
                                'shares': pos.shares,
                                'pct': pos.shares / prev_total,  # Same as cap table %
                                'share_class_id': pos.share_class_id,
                            })
                    break

            # Only show pro rata if there are previous holders
            if prev_snapshot_holders:
                pro_rata_end_row = self._render_pro_rata_editor(
                    sheet, col_map, pref_class_ids, pro_rata_round_id,
                    prev_snapshot_holders, pro_rata_end_row, pref_start_row,
                    prev_label
                )

        # Add Secondary Transactions box below Pro Rata editor (only if there are secondary transactions)
        if secondary_transactions:
            self._render_secondary_editor(
                sheet, col_map, pref_class_ids, secondary_transactions,
                pro_rata_end_row, pref_start_row
            )

        # Add named ranges for key cells to make formulas readable
        self._add_named_ranges_for_snapshot(wb, sheet, snap_cfg.label, pref_class_ids, col_map, totals_row, pre_row, price_row, post_row)

        # Add box around entire cap table
        # Determine bounds: from row 3 (headers) to last valuation row, from column A to last column (% FD)
        first_row = 3
        last_row = post_row
        first_col = 1  # A
        last_col = self._col_to_index(col_map['pct_fd'])

        # Apply thick border to outer edges
        for c in range(first_col, last_col + 1):
            # Top border on header row
            cell = sheet.cell(row=first_row, column=c)
            cell.border = Border(
                left=cell.border.left if cell.border else None,
                right=cell.border.right if cell.border else None,
                top=Side(style='medium'),
                bottom=cell.border.bottom if cell.border else None
            )
            # Bottom border on last row
            cell = sheet.cell(row=last_row, column=c)
            cell.border = Border(
                left=cell.border.left if cell.border else None,
                right=cell.border.right if cell.border else None,
                top=cell.border.top if cell.border else None,
                bottom=Side(style='medium')
            )

        for r in range(first_row, last_row + 1):
            # Left border on column A
            cell = sheet.cell(row=r, column=first_col)
            cell.border = Border(
                left=Side(style='medium'),
                right=cell.border.right if cell.border else None,
                top=cell.border.top if cell.border else None,
                bottom=cell.border.bottom if cell.border else None
            )
            # Right border on last column
            cell = sheet.cell(row=r, column=last_col)
            cell.border = Border(
                left=cell.border.left if cell.border else None,
                right=Side(style='medium'),
                top=cell.border.top if cell.border else None,
                bottom=cell.border.bottom if cell.border else None
            )

        # Store col_map for this sheet so it can be referenced by later sheets
        self._sheet_col_maps[snap_cfg.label] = col_map


    # ------------------------------------------------------------------ #
    def _add_named_ranges_for_snapshot(
        self,
        workbook: Workbook,
        sheet,
        sheet_name: str,
        pref_class_ids: List[str],
        col_map: Dict[str, str],
        totals_row: int,
        pre_row: int,
        price_row: int,
        post_row: int
    ) -> None:
        """Add named ranges to make formulas more readable.

        Creates readable names like:
        - Seed_PreMoney, SeriesA_PreMoney
        - Seed_PPS, SeriesA_PPS
        - Seed_TotalShares, SeriesA_TotalShares
        """
        for pref_id in pref_class_ids:
            # Clean pref_id for use in name (remove spaces, special chars)
            clean_id = pref_id.replace(" ", "").replace("-", "")

            invest_col = col_map.get(f"{pref_id}_invested")
            shares_col = col_map.get(f"{pref_id}_shares")

            if invest_col:
                # Pre-money valuation
                self._add_named_range(
                    workbook, sheet_name,
                    f"{clean_id}_PreMoney",
                    f"{invest_col}{pre_row}"
                )

                # Price per share
                self._add_named_range(
                    workbook, sheet_name,
                    f"{clean_id}_PPS",
                    f"{invest_col}{price_row}"
                )

                # Post-money valuation
                self._add_named_range(
                    workbook, sheet_name,
                    f"{clean_id}_PostMoney",
                    f"{invest_col}{post_row}"
                )

                # Total investment
                self._add_named_range(
                    workbook, sheet_name,
                    f"{clean_id}_TotalInvestment",
                    f"{invest_col}{totals_row}"
                )

            if shares_col:
                # Total shares for this class
                self._add_named_range(
                    workbook, sheet_name,
                    f"{clean_id}_TotalShares",
                    f"{shares_col}{totals_row}"
                )

    # ------------------------------------------------------------------ #
    # Option Pool Editor
    # ------------------------------------------------------------------ #

    def _render_option_pool_editor(
        self,
        sheet,
        col_map: Dict[str, str],
        pref_class_ids: List[str],
        rounds_with_pools: set,
        allocated_row: int,
        available_row: int,
        prev_pref_class_ids: set,
        prev_label: Optional[str],
        totals_row: int
    ) -> None:
        """Render a simple option pool % input cell to the right of the table.

        The input cell allows users to set the option pool as a % of total shares.
        The main table option pool cells are calculated from this percentage.
        """
        # Position: 2 columns after pct_fd, on the ESOP Available row
        editor_col = self._col_to_index(col_map['pct_fd']) + 2
        editor_col_letter = self._col_letter(editor_col)
        input_row = available_row

        # Set column width
        sheet.column_dimensions[editor_col_letter].width = 12

        # Single input cell for option pool %
        input_cell = sheet.cell(row=input_row, column=editor_col)
        input_cell.value = 0.10  # Default 10%
        input_cell.number_format = '0%'
        self._mark_as_input_cell(input_cell, "Option Pool %")

        # Store the input cell reference
        input_ref = f"{editor_col_letter}{input_row}"

        # Update main table option pool cells to use formula based on %
        # Option Pool Shares = Total Shares * Pool %
        # But we need to avoid circular reference since total shares includes option pool
        # Use: Pool = (Other Shares) * Pool% / (1 - Pool%)
        for pref_id in pref_class_ids:
            # Only process rounds that have option pools
            if pref_id not in rounds_with_pools:
                continue

            opt_col = col_map[f"{pref_id}_option_pool"]
            is_prev_round = pref_id in prev_pref_class_ids

            # Allocated row - set to 0 or reference previous
            alloc_cell = sheet[f"{opt_col}{allocated_row}"]
            if is_prev_round and prev_label:
                prev_sheet_ref = f"'{prev_label}'!{opt_col}{allocated_row}"
                alloc_cell.value = f"={prev_sheet_ref}"
                alloc_cell.font = self.green_font
            else:
                alloc_cell.value = 0
                alloc_cell.font = self.black_font
            alloc_cell.number_format = '#,##0'

            # Available row - calculate from % input
            avail_cell = sheet[f"{opt_col}{available_row}"]
            if is_prev_round and prev_label:
                prev_sheet_ref = f"'{prev_label}'!{opt_col}{available_row}"
                avail_cell.value = f"={prev_sheet_ref}"
                avail_cell.font = self.green_font
            else:
                # Only the latest round gets the pool calculation
                # Formula: Pool = (Total - Pool) * %  =>  Pool = Total * % / (1 + %)
                # Simpler: reference total shares excluding this pool column
                other_cols = [col_map["common_shares"]] + [col_map[f"{pid}_shares"] for pid in pref_class_ids]
                other_cols += [col_map[f"{pid}_option_pool"] for pid in pref_class_ids if pid in rounds_with_pools and pid != pref_id]
                other_sum = "+".join(f"{c}{totals_row}" for c in other_cols)
                avail_cell.value = f"=ROUND(({other_sum})*{input_ref}/(1-{input_ref}),0)"
                avail_cell.font = self.black_font
            avail_cell.number_format = '#,##0'

    # ------------------------------------------------------------------ #
    # SAFE Editor
    # ------------------------------------------------------------------ #

    def _render_safe_editor(
        self,
        sheet,
        col_map: Dict[str, str],
        pref_class_ids: List[str],
        safe_rounds: Dict[str, dict],
        post_row: int,
        start_shares_cells: Dict[str, str],
        pps_cells: Dict[str, str],
        prev_pref_class_ids: set,
        prev_label: Optional[str]
    ) -> int:
        """Render a SAFE editor box below the cap table.

        Creates input cells for each SAFE's valuation cap and discount rate.
        The main table PPS cells are formulas referencing these inputs.

        Returns:
            The last row used by the SAFE editor box.
        """
        # Position: below the cap table, starting 2 rows after post_row
        start_row = post_row + 2

        # Header row
        header_cell = sheet.cell(row=start_row, column=1)
        header_cell.value = "SAFE Parameters"
        header_cell.font = Font(bold=True, size=11)
        header_cell.fill = self.calculator_fill

        # Column headers
        label_col = 1
        cap_label_col = 2
        discount_label_col = 3

        sheet.cell(row=start_row, column=cap_label_col).value = "Valuation Cap"
        sheet.cell(row=start_row, column=cap_label_col).font = Font(bold=True)
        sheet.cell(row=start_row, column=cap_label_col).fill = self.calculator_fill

        sheet.cell(row=start_row, column=discount_label_col).value = "Discount %"
        sheet.cell(row=start_row, column=discount_label_col).font = Font(bold=True)
        sheet.cell(row=start_row, column=discount_label_col).fill = self.calculator_fill

        # Set column widths
        sheet.column_dimensions[self._col_letter(cap_label_col)].width = 15
        sheet.column_dimensions[self._col_letter(discount_label_col)].width = 12

        current_row = start_row + 1

        # One row per SAFE with cap and discount inputs
        for pref_id in pref_class_ids:
            if pref_id not in safe_rounds:
                continue

            safe_info = safe_rounds[pref_id]
            valuation_cap = safe_info['valuation_cap']
            discount_rate = safe_info['discount_rate']
            is_prev_round = pref_id in prev_pref_class_ids

            # Label
            label_cell = sheet.cell(row=current_row, column=label_col)
            label_cell.value = pref_id
            label_cell.font = Font(italic=True)
            label_cell.fill = self.calculator_fill

            # Cap input
            cap_cell = sheet.cell(row=current_row, column=cap_label_col)
            cap_cell_ref = f"{self._col_letter(cap_label_col)}{current_row}"
            if valuation_cap:
                cap_cell.value = float(valuation_cap)
                if is_prev_round:
                    cap_cell.font = self.green_font
                else:
                    self._mark_as_input_cell(cap_cell, f"SAFE valuation cap for {pref_id}")
            else:
                cap_cell.value = None
            cap_cell.number_format = '$#,##0'

            # Discount input
            discount_cell = sheet.cell(row=current_row, column=discount_label_col)
            discount_cell_ref = f"{self._col_letter(discount_label_col)}{current_row}"
            if discount_rate:
                discount_cell.value = float(discount_rate)
                if is_prev_round:
                    discount_cell.font = self.green_font
                else:
                    self._mark_as_input_cell(discount_cell, f"SAFE discount for {pref_id} (e.g., 0.20 = 20%)")
            else:
                discount_cell.value = None
            discount_cell.number_format = '0%'

            # Update the main table PPS cell with formula referencing these inputs
            # Skip for previous rounds - they already reference the previous sheet's PPS
            if pref_id in pps_cells and pref_id in start_shares_cells and not is_prev_round:
                pps_cell_ref = pps_cells[pref_id]
                start_cell_ref = start_shares_cells[pref_id]
                pps_cell = sheet[pps_cell_ref]

                if valuation_cap and discount_rate:
                    # Both: MIN of cap-based and discount-based (discount needs priced round PPS)
                    # For now just use cap: =Cap/Shares
                    pps_cell.value = f"=IFERROR({cap_cell_ref}/{start_cell_ref},\"\")"
                    pps_cell.comment = Comment(
                        f"SAFE conversion price for {pref_id}\n"
                        "= Valuation Cap / Starting Shares\n"
                        "(Or use discount if lower: PPS * (1 - Discount))",
                        "System"
                    )
                elif valuation_cap:
                    # Cap only: =Cap/Shares
                    pps_cell.value = f"=IFERROR({cap_cell_ref}/{start_cell_ref},\"\")"
                    pps_cell.comment = Comment(
                        f"SAFE conversion price for {pref_id}\n= Valuation Cap / Starting Shares",
                        "System"
                    )
                elif discount_rate:
                    # Discount only: need priced round PPS reference
                    # Find the next priced round PPS
                    next_priced_pps = None
                    for pid in pref_class_ids:
                        if pid not in safe_rounds and pid in pps_cells:
                            next_priced_pps = pps_cells[pid]
                            break
                    if next_priced_pps:
                        pps_cell.value = f"=IFERROR({next_priced_pps}*(1-{discount_cell_ref}),\"\")"
                        pps_cell.comment = Comment(
                            f"SAFE conversion price for {pref_id}\n= Priced Round PPS * (1 - Discount)",
                            "System"
                        )
                    else:
                        pps_cell.value = None
                        pps_cell.comment = Comment("Discount-based SAFE - no priced round PPS found", "System")

            current_row += 1

        # Add border around SAFE editor box
        last_row = current_row - 1
        for r in range(start_row, last_row + 1):
            for c in range(label_col, discount_label_col + 1):
                cell = sheet.cell(row=r, column=c)
                cell.border = Border(
                    left=Side(style='thin') if c == label_col else None,
                    right=Side(style='thin') if c == discount_label_col else None,
                    top=Side(style='thin') if r == start_row else None,
                    bottom=Side(style='thin') if r == last_row else None
                )

        return last_row

    # ------------------------------------------------------------------ #
    # Pro Rata Editor
    # ------------------------------------------------------------------ #

    def _render_pro_rata_editor(
        self,
        sheet,
        col_map: Dict[str, str],
        pref_class_ids: List[str],
        target_round_id: str,
        prev_snapshot_holders: List[dict],
        prev_editor_end_row: int,
        pref_start_row: int,
        prev_label: str,
    ) -> int:
        """Render a Pro Rata editor box for the target round.

        Shows existing investors' pro rata rights and allows editing their participation.

        Args:
            sheet: The worksheet object
            col_map: Column letter mapping
            pref_class_ids: List of preferred class IDs
            target_round_id: The round ID where pro rata applies
            prev_snapshot_holders: List of {holder_id, shares, pct} from previous snapshot
            prev_editor_end_row: Row after previous editor
            pref_start_row: Start row of preferred holder rows
            prev_label: Label of the previous snapshot sheet (for cross-sheet references)

        Returns:
            The last row used by the pro rata editor box.
        """
        # Position: below the previous editor, starting 2 rows after
        start_row = prev_editor_end_row + 2

        # Column layout
        label_col = 1
        holder_col = 2
        current_pct_col = 3
        pro_rata_col = 4
        participating_col = 5
        investment_col = 6

        # Get previous sheet's pct_fd column for cross-sheet reference
        prev_col_map = self._sheet_col_maps.get(prev_label, {})
        prev_pct_col = prev_col_map.get('pct_fd', 'H')  # Default to H if not found

        # Header row
        header_cell = sheet.cell(row=start_row, column=label_col)
        header_cell.value = f"Pro Rata Rights ({target_round_id})"
        header_cell.font = Font(bold=True, size=11)
        header_cell.fill = self.calculator_fill

        # Total Round Size input (separate row above the table)
        total_round_row = start_row
        total_label = sheet.cell(row=total_round_row, column=holder_col)
        total_label.value = "Total Round Size:"
        total_label.font = Font(bold=True)
        total_label.fill = self.calculator_fill

        total_input = sheet.cell(row=total_round_row, column=current_pct_col)
        total_input_ref = f"{self._col_letter(current_pct_col)}{total_round_row}"
        total_input.value = None  # User inputs total round
        total_input.number_format = '$#,##0'
        self._mark_as_input_cell(total_input, "Total round size (all investors combined)")

        # Column headers
        col_header_row = start_row + 1
        headers = [
            (holder_col, "Investor"),
            (current_pct_col, "Current %"),
            (pro_rata_col, "Pro Rata $"),
            (participating_col, "Participating"),
            (investment_col, "Investment"),
        ]
        for col, label in headers:
            cell = sheet.cell(row=col_header_row, column=col)
            cell.value = label
            cell.font = Font(bold=True)
            cell.fill = self.calculator_fill

        # Set column widths
        sheet.column_dimensions[self._col_letter(holder_col)].width = 20
        sheet.column_dimensions[self._col_letter(current_pct_col)].width = 12
        sheet.column_dimensions[self._col_letter(pro_rata_col)].width = 14
        sheet.column_dimensions[self._col_letter(participating_col)].width = 12
        sheet.column_dimensions[self._col_letter(investment_col)].width = 14

        current_row = col_header_row + 1

        # Store cell references for main table updates
        pro_rata_refs: Dict[str, str] = {}  # holder_id -> investment cell ref

        # One row per existing holder with pro rata rights
        for holder_info in prev_snapshot_holders:
            holder_id = holder_info['holder_id']
            holder_pct = holder_info['pct']

            # Skip if holder has no ownership (shouldn't happen but be safe)
            if holder_pct <= 0:
                continue

            # Holder name
            holder_cell = sheet.cell(row=current_row, column=holder_col)
            holder_cell_ref = f"{self._col_letter(holder_col)}{current_row}"
            holder_cell.value = holder_id
            holder_cell.font = Font(italic=True)
            holder_cell.fill = self.calculator_fill

            # Current % (from previous snapshot - use SUMIF to reference actual ownership)
            pct_cell = sheet.cell(row=current_row, column=current_pct_col)
            pct_cell_ref = f"{self._col_letter(current_pct_col)}{current_row}"
            # Formula: look up holder name in previous sheet's column A, return their % FD
            pct_cell.value = f"=SUMIF('{prev_label}'!$A:$A,{holder_cell_ref},'{prev_label}'!${prev_pct_col}:${prev_pct_col})"
            pct_cell.number_format = '0.0%'
            pct_cell.font = self.green_font  # From previous snapshot
            pct_cell.fill = self.calculator_fill

            # Pro Rata allocation (formula: Current % × Total Round)
            pro_rata_cell = sheet.cell(row=current_row, column=pro_rata_col)
            pro_rata_cell_ref = f"{self._col_letter(pro_rata_col)}{current_row}"
            pro_rata_cell.value = f"=IFERROR({pct_cell_ref}*{total_input_ref},0)"
            pro_rata_cell.number_format = '$#,##0'
            pro_rata_cell.font = self.black_font  # Calculated
            pro_rata_cell.fill = self.calculator_fill

            # Participating (dropdown: Yes/No or %)
            participating_cell = sheet.cell(row=current_row, column=participating_col)
            participating_cell_ref = f"{self._col_letter(participating_col)}{current_row}"
            participating_cell.value = 1.0  # Default: 100% participating
            participating_cell.number_format = '0%'
            self._mark_as_input_cell(participating_cell, "Participation rate (100% = full pro rata, 0% = not participating)")

            # Investment (formula: Pro Rata × Participation %)
            invest_cell = sheet.cell(row=current_row, column=investment_col)
            invest_cell_ref = f"{self._col_letter(investment_col)}{current_row}"
            invest_cell.value = f"=IFERROR({pro_rata_cell_ref}*{participating_cell_ref},0)"
            invest_cell.number_format = '$#,##0'
            invest_cell.font = self.black_font  # Calculated from participation
            invest_cell.fill = self.calculator_fill

            # Store reference for main table
            pro_rata_refs[holder_id] = invest_cell_ref

            current_row += 1

        # Add "New Lead Investor" row at the bottom
        # This is for the new investor who takes the remainder
        lead_row = current_row
        lead_label = sheet.cell(row=lead_row, column=holder_col)
        lead_label.value = "(New Lead)"
        lead_label.font = Font(italic=True, bold=True)
        lead_label.fill = self.calculator_fill

        # Lead gets: Total Round - Sum of all pro rata investments
        lead_pct_cell = sheet.cell(row=lead_row, column=current_pct_col)
        lead_pct_cell.value = None  # No current ownership
        lead_pct_cell.fill = self.calculator_fill

        lead_pro_rata = sheet.cell(row=lead_row, column=pro_rata_col)
        lead_pro_rata.value = None  # N/A
        lead_pro_rata.fill = self.calculator_fill

        lead_participating = sheet.cell(row=lead_row, column=participating_col)
        lead_participating.value = None  # N/A
        lead_participating.fill = self.calculator_fill

        # Lead investment = Total - Sum(pro rata investments)
        lead_invest = sheet.cell(row=lead_row, column=investment_col)
        if pro_rata_refs:
            pro_rata_sum = "+".join(pro_rata_refs.values())
            lead_invest.value = f"=IFERROR({total_input_ref}-({pro_rata_sum}),0)"
        else:
            lead_invest.value = f"={total_input_ref}"
        lead_invest.number_format = '$#,##0'
        lead_invest.font = self.black_font
        lead_invest.fill = self.calculator_fill

        current_row += 1

        # Totals row
        totals_label = sheet.cell(row=current_row, column=holder_col)
        totals_label.value = "Total"
        totals_label.font = Font(bold=True)
        totals_label.fill = self.calculator_fill

        # Skip pct/pro rata columns for totals
        for col in [current_pct_col, pro_rata_col, participating_col]:
            cell = sheet.cell(row=current_row, column=col)
            cell.fill = self.calculator_fill

        # Total investments should equal Total Round
        totals_invest = sheet.cell(row=current_row, column=investment_col)
        invest_range = f"{self._col_letter(investment_col)}{col_header_row+1}:{self._col_letter(investment_col)}{current_row-1}"
        totals_invest.value = f"=SUM({invest_range})"
        totals_invest.number_format = '$#,##0'
        totals_invest.font = Font(bold=True)
        totals_invest.fill = self.calculator_fill

        current_row += 1

        # Add border around pro rata editor box
        last_row = current_row - 1
        last_col = investment_col
        for r in range(start_row, last_row + 1):
            for c in range(label_col, last_col + 1):
                cell = sheet.cell(row=r, column=c)
                current_border = cell.border if cell.border else Border()
                cell.border = Border(
                    left=Side(style='thin') if c == label_col else current_border.left,
                    right=Side(style='thin') if c == last_col else current_border.right,
                    top=Side(style='thin') if r == start_row else current_border.top,
                    bottom=Side(style='thin') if r == last_row else current_border.bottom
                )

        # Update main table investment cells to reference this pro rata editor
        if target_round_id in pref_class_ids:
            invest_col = col_map[f"{target_round_id}_invested"]

            # Find holder rows in main table and update their investment formulas
            row = pref_start_row
            while True:
                holder_cell = sheet[f"A{row}"]
                if holder_cell.value is None or holder_cell.value == "":
                    break
                if holder_cell.value in ["Totals", "Starting Shares (pre-round)", "Pre-Money Valuation",
                                         "Price per Share", "Post-Money Valuation", "Preferred Rounds",
                                         "Allocated Options", "ESOP Available"]:
                    row += 1
                    continue

                holder_id = holder_cell.value
                invest_cell = sheet[f"{invest_col}{row}"]

                # If this holder has a pro rata entry, reference it
                if holder_id in pro_rata_refs:
                    invest_cell.value = f"={pro_rata_refs[holder_id]}"
                    invest_cell.font = self.black_font  # Formula reference
                    invest_cell.number_format = '$#,##0'

                row += 1

        return last_row

    # ------------------------------------------------------------------ #
    # Secondary Transactions Editor
    # ------------------------------------------------------------------ #

    def _render_secondary_editor(
        self,
        sheet,
        col_map: Dict[str, str],
        pref_class_ids: List[str],
        secondary_transactions: List[dict],
        prev_editor_end_row: int,
        pref_start_row: int,
    ) -> int:
        """Render a Secondary Transactions editor box below the SAFE editor.

        Creates input cells for each secondary transaction with seller, buyer,
        share class, shares, and price. The main table share cells are updated
        to reference these inputs via SUMIF formulas.

        Returns:
            The last row used by the secondary editor box.
        """
        # Position: below the previous editor, starting 2 rows after
        start_row = prev_editor_end_row + 2

        # Check if any transactions have alchemy (buyer gets different class)
        has_alchemy = any(txn['share_class'] != txn['resulting_class'] for txn in secondary_transactions)

        # Column layout for secondary transactions box
        # Shares is now a formula = Total $ / Price, so Total comes before Shares
        seller_col = 1
        buyer_col = 2
        seller_class_col = 3   # Seller's class (what they give up)
        buyer_class_col = 4 if has_alchemy else None  # Buyer's class (alchemy)
        total_col = 5 if has_alchemy else 4      # Input: $ amount paid for secondary
        price_col = 6 if has_alchemy else 5      # Input: secondary price per share
        shares_col = 7 if has_alchemy else 6     # Formula: Total $ / Price

        # Header row
        header_cell = sheet.cell(row=start_row, column=seller_col)
        header_cell.value = "Secondary Transactions" + (" (with Alchemy)" if has_alchemy else "")
        header_cell.font = Font(bold=True, size=11)
        header_cell.fill = self.calculator_fill

        # Column headers
        headers = [
            (seller_col, "Seller"),
            (buyer_col, "Buyer"),
            (seller_class_col, "Sells Class"),
        ]
        if has_alchemy:
            headers.append((buyer_class_col, "Gets Class"))
        headers.extend([
            (total_col, "Total $"),
            (price_col, "Price/Share"),
            (shares_col, "Shares"),
        ])
        for col, label in headers:
            cell = sheet.cell(row=start_row, column=col)
            cell.value = label
            cell.font = Font(bold=True)
            cell.fill = self.calculator_fill

        # Set column widths
        sheet.column_dimensions[self._col_letter(seller_col)].width = 18
        sheet.column_dimensions[self._col_letter(buyer_col)].width = 18
        sheet.column_dimensions[self._col_letter(seller_class_col)].width = 12
        if has_alchemy:
            sheet.column_dimensions[self._col_letter(buyer_class_col)].width = 12
        sheet.column_dimensions[self._col_letter(total_col)].width = 14
        sheet.column_dimensions[self._col_letter(price_col)].width = 12
        sheet.column_dimensions[self._col_letter(shares_col)].width = 12

        current_row = start_row + 1

        # Store cell references for each transaction (for main table formulas)
        secondary_cell_refs: List[dict] = []

        # One row per secondary transaction
        for txn in secondary_transactions:
            # Seller (input)
            seller_cell = sheet.cell(row=current_row, column=seller_col)
            seller_cell.value = txn['from_holder']
            self._mark_as_input_cell(seller_cell, "Seller (who is selling shares)")

            # Buyer (input)
            buyer_cell = sheet.cell(row=current_row, column=buyer_col)
            buyer_cell.value = txn['to_holder']
            self._mark_as_input_cell(buyer_cell, "Buyer (who is acquiring shares)")

            # Seller's Share Class (what they give up)
            seller_class_cell = sheet.cell(row=current_row, column=seller_class_col)
            seller_class_cell.value = txn['share_class']
            self._mark_as_input_cell(seller_class_cell, "Share class seller gives up")

            # Buyer's Share Class (what they receive - for alchemy)
            if has_alchemy and buyer_class_col:
                buyer_class_cell = sheet.cell(row=current_row, column=buyer_class_col)
                buyer_class_cell.value = txn['resulting_class']
                self._mark_as_input_cell(buyer_class_cell, "Share class buyer receives (alchemy)")

            # Total $ (input: amount paid for secondary)
            total_cell = sheet.cell(row=current_row, column=total_col)
            total_cell_ref = f"{self._col_letter(total_col)}{current_row}"
            # Calculate total from shares × price
            if txn['price_per_share'] and txn['shares']:
                total_cell.value = float(txn['shares'] * txn['price_per_share'])
            else:
                total_cell.value = float(txn['shares']) if txn['shares'] else None
            total_cell.number_format = '$#,##0'
            self._mark_as_input_cell(total_cell, "Total $ paid for this secondary transaction")

            # Price per Share (input: secondary transaction price)
            price_cell = sheet.cell(row=current_row, column=price_col)
            price_cell_ref = f"{self._col_letter(price_col)}{current_row}"
            if txn['price_per_share']:
                price_cell.value = float(txn['price_per_share'])
            else:
                price_cell.value = None
            price_cell.number_format = '$0.00'
            self._mark_as_input_cell(price_cell, "Price per share for this secondary transaction")

            # Shares (formula: Total $ / Price)
            shares_cell = sheet.cell(row=current_row, column=shares_col)
            shares_cell_ref = f"{self._col_letter(shares_col)}{current_row}"
            shares_cell.value = f"=IFERROR({total_cell_ref}/{price_cell_ref},0)"
            shares_cell.font = self.black_font
            shares_cell.number_format = '#,##0'

            # Store references for main table updates
            secondary_cell_refs.append({
                'row': current_row,
                'seller_ref': f"{self._col_letter(seller_col)}{current_row}",
                'buyer_ref': f"{self._col_letter(buyer_col)}{current_row}",
                'seller_class_ref': f"{self._col_letter(seller_class_col)}{current_row}",
                'shares_ref': shares_cell_ref,
                'from_holder': txn['from_holder'],
                'to_holder': txn['to_holder'],
                'share_class': txn['share_class'],        # Seller's class
                'resulting_class': txn['resulting_class'],  # Buyer's class (may differ)
            })

            current_row += 1

        # Add border around secondary editor box
        last_row = current_row - 1
        for r in range(start_row, last_row + 1):
            for c in range(seller_col, total_col + 1):
                cell = sheet.cell(row=r, column=c)
                cell.border = Border(
                    left=Side(style='thin') if c == seller_col else None,
                    right=Side(style='thin') if c == total_col else None,
                    top=Side(style='thin') if r == start_row else None,
                    bottom=Side(style='thin') if r == last_row else None
                )

        # Now update the main table share cells to reference the secondary box
        # For each holder row, if they are involved in secondary:
        #   - Sellers: Current Shares = Original - SUMIF(sellers match, class matches)
        #   - Buyers: Current Shares = Primary + SUMIF(buyers match, class matches)
        self._update_main_table_for_secondary(
            sheet, col_map, pref_class_ids, pref_start_row,
            secondary_cell_refs
        )

        return last_row

    def _update_main_table_for_secondary(
        self,
        sheet,
        col_map: Dict[str, str],
        pref_class_ids: List[str],
        pref_start_row: int,
        secondary_cell_refs: List[dict],
    ) -> None:
        """Update main table share cells to account for secondary transactions.

        Uses SUMIF to dynamically sum shares sold/acquired based on the
        secondary transactions box inputs.
        """
        # Build lookup: (holder_id, share_class) -> list of secondary refs affecting them
        # With alchemy, seller loses from share_class, buyer gains in resulting_class
        holder_secondary: Dict[tuple, List[dict]] = {}
        for ref in secondary_cell_refs:
            # Seller loses shares from their original class
            seller_key = (ref['from_holder'], ref['share_class'])
            if seller_key not in holder_secondary:
                holder_secondary[seller_key] = []
            holder_secondary[seller_key].append({'type': 'sold', **ref})

            # Buyer gains shares in the resulting class (may differ with alchemy)
            buyer_class = ref.get('resulting_class', ref['share_class'])
            buyer_key = (ref['to_holder'], buyer_class)
            if buyer_key not in holder_secondary:
                holder_secondary[buyer_key] = []
            holder_secondary[buyer_key].append({'type': 'acquired', **ref})

        # Iterate through the main table preferred holder rows
        for pref_id in pref_class_ids:
            shares_col = col_map[f"{pref_id}_shares"]

            # Find rows with this share class
            # We need to scan the main table to find holder rows
            # Start from pref_start_row and look for rows with data
            row = pref_start_row
            while True:
                holder_cell = sheet[f"A{row}"]
                if holder_cell.value is None or holder_cell.value == "":
                    break
                if holder_cell.value in ["Totals", "Starting Shares (pre-round)", "Pre-Money Valuation",
                                         "Price per Share", "Post-Money Valuation", "Preferred Rounds",
                                         "Allocated Options", "ESOP Available"]:
                    row += 1
                    continue

                holder_id = holder_cell.value
                shares_cell = sheet[f"{shares_col}{row}"]

                # Check if this holder+class combination has secondary transactions
                key = (holder_id, pref_id)
                if key in holder_secondary:
                    txns = holder_secondary[key]

                    # Get the current formula/value
                    current_value = shares_cell.value

                    # Build adjustment formula
                    # Sellers: subtract shares sold
                    # Buyers: add shares acquired
                    sold_sum = []
                    acquired_sum = []

                    for txn in txns:
                        if txn['type'] == 'sold':
                            # SUMIF(seller_range, holder_id, shares_range) where class matches
                            # Since class is fixed per column, we just need to match seller
                            sold_sum.append(txn['shares_ref'])
                        else:  # acquired
                            acquired_sum.append(txn['shares_ref'])

                    # Modify the cell formula
                    if current_value and str(current_value).startswith('='):
                        # Current value is a formula, append to it
                        base_formula = str(current_value)
                        if sold_sum:
                            base_formula += f"-{'+'.join(sold_sum)}"
                        if acquired_sum:
                            base_formula += f"+{'+'.join(acquired_sum)}"
                        shares_cell.value = base_formula
                    else:
                        # Current value is hardcoded, wrap it
                        base_value = current_value if current_value else 0
                        formula_parts = [str(base_value)]
                        if sold_sum:
                            formula_parts.append(f"-({'+'.join(sold_sum)})")
                        if acquired_sum:
                            formula_parts.append(f"+({'+'.join(acquired_sum)})")
                        shares_cell.value = f"={''.join(formula_parts)}"

                row += 1

    # ------------------------------------------------------------------ #
    # Interactive Features (Phase 1: Essential)
    # ------------------------------------------------------------------ #

    def _mark_as_input_cell(self, cell, description: Optional[str] = None) -> None:
        """Mark a cell as user-editable input with visual formatting.

        Applies:
        - Light blue background
        - Bold text
        - Medium blue border
        - Optional comment with guidance

        Args:
            cell: The openpyxl cell object
            description: Optional guidance text to add as comment
        """
        cell.fill = self.input_cell_fill
        cell.font = self.input_cell_font
        cell.border = self.input_cell_border

        if description:
            cell.comment = Comment(description, "Cap Table Generator")

    def _add_dropdown_validation(
        self,
        worksheet,
        cell_ref: str,
        options: List[str],
        prompt_title: str = "Select Value",
        prompt_text: str = "Choose from the dropdown"
    ) -> None:
        """Add data validation dropdown to a cell.

        Args:
            worksheet: The worksheet object
            cell_ref: Cell reference (e.g., "B36")
            options: List of dropdown options
            prompt_title: Title for validation prompt
            prompt_text: Description text for validation prompt
        """
        dv = DataValidation(
            type="list",
            formula1=f'"{",".join(options)}"',
            allow_blank=False
        )
        dv.prompt = prompt_text
        dv.promptTitle = prompt_title
        worksheet.add_data_validation(dv)
        dv.add(cell_ref)

    def _add_named_range(
        self,
        workbook: Workbook,
        sheet_name: str,
        name: str,
        cell_ref: str
    ) -> None:
        """Add a named range for better formula readability.

        Args:
            workbook: The workbook object
            sheet_name: Sheet name
            name: Name for the range (e.g., "TargetPoolPct")
            cell_ref: Cell reference (e.g., "B36")
        """
        workbook.defined_names.add(
            DefinedName(
                name=name,
                attr_text=f"'{sheet_name}'!${cell_ref}"
            )
        )

    @staticmethod
    def _col_to_index(col_letter: str) -> int:
        """Convert Excel column letter to 1-based index."""
        idx = 0
        for char in col_letter:
            idx = idx * 26 + (ord(char) - 64)
        return idx

    # ------------------------------------------------------------------ #
    @staticmethod
    def _col_letter(idx: int) -> str:
        """Convert 1-based column index to Excel column letter."""
        letter = ""
        while idx > 0:
            idx, rem = divmod(idx - 1, 26)
            letter = chr(65 + rem) + letter
        return letter


__all__ = ["RoundSheetRenderer"]
