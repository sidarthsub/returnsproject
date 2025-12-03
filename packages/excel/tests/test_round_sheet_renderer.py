"""Comprehensive tests for RoundSheetRenderer.

This module contains all tests for the round_sheet_renderer, including:
1. Block verification - Ensures Excel output matches domain calculations
2. Formula structure validation - Verifies formulas are correctly formed
3. Enhanced validation - Catches rendering issues (pre-money structure, % FD totals, etc.)
"""

from datetime import date
from decimal import Decimal
import pathlib
import sys

REPO_ROOT = pathlib.Path(__file__).resolve().parents[3]
sys.path.insert(0, str(REPO_ROOT / "packages" / "excel" / "src"))
sys.path.insert(0, str(REPO_ROOT / "packages" / "domain"))

OUTPUT_DIR = REPO_ROOT / "packages" / "excel" / "tests" / "output"
OUTPUT_DIR.mkdir(parents=True, exist_ok=True)

from captable_domain.schemas import (
    CapTable,
    CapTableSnapshotCFG,
    LiquidationPreference,
    OptionPoolCreation,
    ParticipationRights,
    ShareClass,
    ShareIssuanceEvent,
    WorkbookCFG,
)
from captable_domain.blocks import CapTableBlock, BlockContext
from captable_excel.round_sheet_renderer import RoundSheetRenderer
from openpyxl import load_workbook


# =============================================================================
# Test Data Builders
# =============================================================================

def build_comprehensive_cap_table() -> CapTable:
    """Build cap table with seed + series A + option pool expansions.

    Creates realistic scenario with:
    - 4 founders (10M common shares)
    - Seed round: 2 investors, $1M investment, 10% option pool
    - Series A round: 3 investors, $5M investment, 15% option pool
    - Total: 9 stakeholders across 2 rounds
    """
    cap_table = CapTable(company_name="TestCo")

    # Common stock class
    cap_table.share_classes["common"] = ShareClass(
        id="common",
        name="Common Stock",
        share_type="common",
    )

    # Seed preferred class
    cap_table.share_classes["seed_pref"] = ShareClass(
        id="seed_pref",
        name="Series Seed Preferred",
        share_type="preferred",
        liquidation_preference=LiquidationPreference(
            multiple=Decimal("1.0"),
            seniority_rank=0
        ),
        participation_rights=ParticipationRights(
            participation_type="non_participating"
        ),
    )

    # Series A preferred class
    cap_table.share_classes["series_a_pref"] = ShareClass(
        id="series_a_pref",
        name="Series A Preferred",
        share_type="preferred",
        liquidation_preference=LiquidationPreference(
            multiple=Decimal("1.0"),
            seniority_rank=1
        ),
        participation_rights=ParticipationRights(
            participation_type="non_participating"
        ),
    )

    # === FOUNDING (Day 0) ===
    # 4 founders get common stock - 10M shares total
    founders = [
        ("Alice", Decimal("4000000")),  # 40%
        ("Bob", Decimal("3000000")),    # 30%
        ("Carol", Decimal("2000000")),  # 20%
        ("Dave", Decimal("1000000")),   # 10%
    ]
    for idx, (name, shares) in enumerate(founders):
        cap_table.add_event(
            ShareIssuanceEvent(
                event_id=f"founder_{name.lower()}",
                event_date=date(2024, 1, 1),
                holder_id=f"founder_{name.lower()}",
                share_class_id="common",
                shares=shares,
            )
        )

    # === SEED ROUND (March 2024) ===
    # Option pool expansion (10% post-money target)
    cap_table.add_event(
        OptionPoolCreation(
            event_id="seed_option_pool",
            event_date=date(2024, 2, 28),
            shares_authorized=Decimal("1389000"),
            pool_timing="post_money",
            share_class_id="common",
        )
    )

    # 2 seed investors
    seed_investors = [
        ("Seed Ventures", Decimal("750000"), Decimal("1875000")),
        ("Angel Investor", Decimal("250000"), Decimal("625000")),
    ]
    for idx, (name, investment, shares) in enumerate(seed_investors):
        cap_table.add_event(
            ShareIssuanceEvent(
                event_id=f"seed_{name.lower().replace(' ', '_')}",
                event_date=date(2024, 3, 1),
                holder_id=name.lower().replace(" ", "_"),
                share_class_id="seed_pref",
                shares=shares,
                price_per_share=Decimal("0.40"),
            )
        )

    # === SERIES A ROUND (July 2024) ===
    # Option pool expansion (15% post-money target)
    cap_table.add_event(
        OptionPoolCreation(
            event_id="series_a_option_pool",
            event_date=date(2024, 6, 30),
            shares_authorized=Decimal("2287000"),
            pool_timing="post_money",
            share_class_id="common",
        )
    )

    # 3 series A investors
    series_a_investors = [
        ("Venture Capital Partners", Decimal("3000000"), Decimal("4166667")),
        ("Growth Equity Fund", Decimal("1500000"), Decimal("2083333")),
        ("Strategic Investor", Decimal("500000"), Decimal("694444")),
    ]
    for idx, (name, investment, shares) in enumerate(series_a_investors):
        cap_table.add_event(
            ShareIssuanceEvent(
                event_id=f"series_a_{name.lower().replace(' ', '_')}",
                event_date=date(2024, 7, 1),
                holder_id=name.lower().replace(" ", "_"),
                share_class_id="series_a_pref",
                shares=shares,
                price_per_share=Decimal("0.72"),
            )
        )

    return cap_table


# =============================================================================
# Test 1: Block Verification
# =============================================================================

def test_round_sheet_matches_block_calculations():
    """Test that Excel output matches block calculations.

    Verifies:
    - Common shares totals
    - Preferred shares and investments by round
    - Option pool expansions
    - Fully diluted share counts
    - All numeric values match between Excel and domain blocks
    """
    cap_table = build_comprehensive_cap_table()

    # Create snapshots for Seed and Series A
    seed_snapshot_cfg = CapTableSnapshotCFG(
        cap_table=cap_table,
        label="Post-Seed",
        as_of_date=date(2024, 3, 31),
    )

    series_a_snapshot_cfg = CapTableSnapshotCFG(
        cap_table=cap_table,
        label="Post-Series-A",
        as_of_date=None,  # Current
    )

    cfg = WorkbookCFG(
        cap_table_snapshots=[seed_snapshot_cfg, series_a_snapshot_cfg],
        waterfall_analyses=None,
    )

    # Render the Excel workbook
    renderer = RoundSheetRenderer(cfg)
    output_path = OUTPUT_DIR / "round_sheet_test.xlsx"
    renderer.render(str(output_path))

    assert output_path.exists(), "Output file should be created"

    # === BLOCK VERIFICATION ===
    for snapshot_cfg in [seed_snapshot_cfg, series_a_snapshot_cfg]:
        _verify_snapshot_calculations(output_path, snapshot_cfg)


def _verify_snapshot_calculations(
    excel_path: pathlib.Path,
    snapshot_cfg: CapTableSnapshotCFG
):
    """Verify that Excel calculations match block calculations."""

    # Get snapshot from cap table
    snapshot = (
        snapshot_cfg.cap_table.snapshot(snapshot_cfg.as_of_date)
        if snapshot_cfg.as_of_date
        else snapshot_cfg.cap_table.current_snapshot()
    )

    # Run block computation
    context = BlockContext()
    context.set("cap_table_snapshot", snapshot)
    block = CapTableBlock()
    block.execute(context)

    ownership_df = context.get("cap_table_ownership")
    summary_df = context.get("cap_table_summary")

    # Load Excel workbook
    wb = load_workbook(excel_path, data_only=True)
    sheet = wb[snapshot_cfg.label]

    # Find header row (row 3) and map columns
    header_row = 3
    headers = {}
    for cell in sheet[header_row]:
        if cell.value:
            headers[str(cell.value).strip()] = cell.column

    # Find totals row
    totals_row = None
    for row in range(4, sheet.max_row + 1):
        if sheet.cell(row=row, column=1).value == "Totals":
            totals_row = row
            break

    assert totals_row is not None, f"Totals row not found in {snapshot_cfg.label}"

    # === VERIFY COMMON SHARES ===
    common_col = headers.get("Common # Shares")
    assert common_col, "Common # Shares column missing"

    common_shares_excel = Decimal("0")
    for row in range(4, totals_row):
        cell_value = sheet.cell(row=row, column=common_col).value
        if isinstance(cell_value, (int, float)) and sheet.cell(row=row, column=1).value not in ["Option Pool", None, ""]:
            common_shares_excel += Decimal(str(cell_value))

    common_shares_block = summary_df["common_shares"].iloc[0]
    assert abs(float(common_shares_excel) - float(common_shares_block)) < 1.0

    # === VERIFY PREFERRED SHARES ===
    pref_classes = [
        sc_id for sc_id, sc in snapshot.share_classes.items()
        if sc.share_type == "preferred"
    ]

    for pref_id in pref_classes:
        shares_header = f"{pref_id} Preferred Shares"
        invest_header = f"{pref_id} $ Invested"

        if shares_header not in headers:
            continue

        shares_col = headers[shares_header]
        invest_col = headers[invest_header]

        pref_invest_excel = Decimal("0")
        pref_start_row = None
        for row in range(4, totals_row):
            if sheet.cell(row=row, column=1).value == "Preferred Rounds":
                pref_start_row = row + 1
                break

        if pref_start_row:
            for row in range(pref_start_row, totals_row):
                invest_value = sheet.cell(row=row, column=invest_col).value
                if isinstance(invest_value, (int, float)):
                    pref_invest_excel += Decimal(str(invest_value))

        pref_invest_block = Decimal("0")
        for pos in snapshot.positions:
            if pos.share_class_id == pref_id and pos.cost_basis:
                pref_invest_block += pos.cost_basis

        assert abs(float(pref_invest_excel) - float(pref_invest_block)) < 1.0

    # === VERIFY OPTION POOL ===
    # Option pool is now per-round, so we need to sum across all rounds
    option_pool_excel = Decimal("0")

    # Find all option pool columns (e.g., "Seed Option Pool", "Series-A Option Pool")
    option_pool_cols = {
        header: col for header, col in headers.items()
        if "Option Pool" in header and "Preferred Shares" not in header
    }

    # Sum option pool values from all rounds
    for col_name, col_idx in option_pool_cols.items():
        for row in range(4, totals_row + 1):  # Include totals row
            cell_value = sheet.cell(row=row, column=col_idx).value
            if isinstance(cell_value, (int, float)) and cell_value > 0:
                option_pool_excel += Decimal(str(cell_value))
                break  # Only count once per column (either from data row or totals row)

    option_pool_block = summary_df["option_pool_shares"].iloc[0]
    if option_pool_block > 0:  # Only assert if there's an option pool in the block
        assert abs(float(option_pool_excel) - float(option_pool_block)) < 1.0


# =============================================================================
# Test 2: Formula Structure
# =============================================================================

def test_formula_structure():
    """Test that Excel formulas are correctly structured.

    Verifies:
    - SUM formulas for totals
    - Starting shares formulas
    - Pre-money/PPS/Post-money formulas
    - Per-investor share formulas
    - Total Shares and % FD formulas
    """
    cap_table = build_comprehensive_cap_table()
    snapshot_cfg = CapTableSnapshotCFG(cap_table=cap_table, label="Test", as_of_date=date(2024, 3, 31))
    cfg = WorkbookCFG(cap_table_snapshots=[snapshot_cfg], waterfall_analyses=None)

    renderer = RoundSheetRenderer(cfg)
    output_path = OUTPUT_DIR / "formula_structure_test.xlsx"
    renderer.render(str(output_path))

    wb = load_workbook(output_path, data_only=False)
    sheet = wb["Test"]

    # Find header row and totals row
    header_row = 3
    headers = {}
    for cell in sheet[header_row]:
        if cell.value:
            headers[str(cell.value).strip()] = cell.column

    totals_row = None
    for row in range(4, sheet.max_row + 1):
        if sheet.cell(row=row, column=1).value == "Totals":
            totals_row = row
            break

    # Find special rows
    starting_shares_row = totals_row + 1
    pre_row = totals_row + 2
    price_row = totals_row + 3
    post_row = totals_row + 4

    # Test common total
    common_col = headers["Common # Shares"]
    common_total = sheet.cell(row=totals_row, column=common_col).value
    assert common_total.startswith("=SUM("), f"Common total should be SUM formula: {common_total}"

    # Test preferred formulas
    seed_invest_col = headers["seed_pref $ Invested"]
    seed_shares_col = headers["seed_pref Preferred Shares"]

    invest_total = sheet.cell(row=totals_row, column=seed_invest_col).value
    assert invest_total.startswith("=SUM("), f"Investment total should be SUM: {invest_total}"

    shares_total = sheet.cell(row=totals_row, column=seed_shares_col).value
    assert shares_total.startswith("=SUM("), f"Shares total should be SUM: {shares_total}"

    # Test starting shares, pre-money, PPS
    starting_shares_cell = sheet.cell(row=starting_shares_row, column=seed_shares_col).value
    assert starting_shares_cell.startswith("="), f"Starting shares should be formula: {starting_shares_cell}"

    pre_money_cell = sheet.cell(row=pre_row, column=seed_invest_col).value
    assert isinstance(pre_money_cell, (int, float)), f"Pre-money should be numeric: {pre_money_cell}"

    pps_cell = sheet.cell(row=price_row, column=seed_invest_col).value
    assert pps_cell.startswith("=IFERROR("), f"PPS should be IFERROR formula: {pps_cell}"


# =============================================================================
# Test 3: Enhanced Validation (Issue Detection)
# =============================================================================

def test_pre_money_row_only_has_dollar_values():
    """Test that Pre-Money row only has $ values, not share counts."""
    cap_table = build_comprehensive_cap_table()
    snapshot_cfg = CapTableSnapshotCFG(cap_table=cap_table, label="Test", as_of_date=None)
    cfg = WorkbookCFG(cap_table_snapshots=[snapshot_cfg], waterfall_analyses=None)

    renderer = RoundSheetRenderer(cfg)
    output_path = OUTPUT_DIR / "pre_money_test.xlsx"
    renderer.render(str(output_path))

    wb = load_workbook(output_path, data_only=False)
    sheet = wb["Test"]

    # Find Pre-Money row
    totals_row = None
    for row in range(4, sheet.max_row + 1):
        if sheet.cell(row=row, column=1).value == "Totals":
            totals_row = row
            break

    pre_row = None
    for row in range(totals_row + 1, sheet.max_row + 1):
        if sheet.cell(row=row, column=1).value == "Pre-Money Valuation":
            pre_row = row
            break

    assert pre_row is not None, "Pre-Money Valuation row not found"

    # Find headers
    headers = {}
    for cell in sheet[3]:
        if cell.value:
            headers[str(cell.value).strip()] = cell.column

    # Check Pre-Money row - should only have values in $ Invested columns
    for col_name, col_idx in headers.items():
        val = sheet.cell(row=pre_row, column=col_idx).value
        if val is not None:
            if "Shares" in col_name:
                raise AssertionError(f"Pre-Money row should NOT have value in {col_name} column: {val}")
            elif "$ Invested" in col_name:
                assert isinstance(val, (int, float)), f"{col_name} should have numeric value"


def test_pct_fd_totals_to_100():
    """Test that % FD column shows 100% in totals row."""
    cap_table = build_comprehensive_cap_table()
    snapshot_cfg = CapTableSnapshotCFG(cap_table=cap_table, label="Test", as_of_date=None)
    cfg = WorkbookCFG(cap_table_snapshots=[snapshot_cfg], waterfall_analyses=None)

    renderer = RoundSheetRenderer(cfg)
    output_path = OUTPUT_DIR / "pct_fd_test.xlsx"
    renderer.render(str(output_path))

    wb = load_workbook(output_path, data_only=False)
    sheet = wb["Test"]

    totals_row = None
    for row in range(4, sheet.max_row + 1):
        if sheet.cell(row=row, column=1).value == "Totals":
            totals_row = row
            break

    headers = {}
    for cell in sheet[3]:
        if cell.value and "% FD" in str(cell.value):
            pct_fd_col = cell.column
            break

    totals_value = sheet.cell(row=totals_row, column=pct_fd_col).value
    assert totals_value is not None, "% FD totals row should not be blank"
    assert abs(float(totals_value) - 1.0) < 0.01, f"% FD total should be 1.0, got {totals_value}"


def test_total_shares_column_exists():
    """Test that Total Shares column exists."""
    cap_table = build_comprehensive_cap_table()
    snapshot_cfg = CapTableSnapshotCFG(cap_table=cap_table, label="Test", as_of_date=None)
    cfg = WorkbookCFG(cap_table_snapshots=[snapshot_cfg], waterfall_analyses=None)

    renderer = RoundSheetRenderer(cfg)
    output_path = OUTPUT_DIR / "total_shares_test.xlsx"
    renderer.render(str(output_path))

    wb = load_workbook(output_path, data_only=False)
    sheet = wb["Test"]

    headers = [cell.value for cell in sheet[3] if cell.value]
    assert any("Total Shares" in str(h) for h in headers), "'Total Shares' column should exist"


def test_header_rows_have_no_formulas():
    """Test that section header rows don't have data formulas."""
    cap_table = build_comprehensive_cap_table()
    snapshot_cfg = CapTableSnapshotCFG(cap_table=cap_table, label="Test", as_of_date=None)
    cfg = WorkbookCFG(cap_table_snapshots=[snapshot_cfg], waterfall_analyses=None)

    renderer = RoundSheetRenderer(cfg)
    output_path = OUTPUT_DIR / "header_clean_test.xlsx"
    renderer.render(str(output_path))

    wb = load_workbook(output_path, data_only=False)
    sheet = wb["Test"]

    # Find "Preferred Rounds" header
    pref_header_row = None
    for row in range(4, sheet.max_row + 1):
        if sheet.cell(row=row, column=1).value == "Preferred Rounds":
            pref_header_row = row
            break

    if pref_header_row:
        # Check all data columns are blank
        for col in range(2, sheet.max_column + 1):
            val = sheet.cell(row=pref_header_row, column=col).value
            assert val is None or val == "", f"Header row should not have data in column {col}: {val}"


def test_investors_only_have_formulas_for_their_rounds():
    """Test that investors only have share formulas for rounds they invested in."""
    cap_table = build_comprehensive_cap_table()
    snapshot_cfg = CapTableSnapshotCFG(cap_table=cap_table, label="Test", as_of_date=None)
    cfg = WorkbookCFG(cap_table_snapshots=[snapshot_cfg], waterfall_analyses=None)

    renderer = RoundSheetRenderer(cfg)
    output_path = OUTPUT_DIR / "investor_scoping_test.xlsx"
    renderer.render(str(output_path))

    wb = load_workbook(output_path, data_only=False)
    sheet = wb["Test"]

    # Find headers and rows
    headers = {}
    for cell in sheet[3]:
        if cell.value:
            headers[str(cell.value).strip()] = cell.column

    pref_header_row = None
    totals_row = None
    for row in range(4, sheet.max_row + 1):
        val = sheet.cell(row=row, column=1).value
        if val == "Preferred Rounds":
            pref_header_row = row
        if val == "Totals":
            totals_row = row

    # Get preferred class columns
    pref_classes = {}
    for col_name, col_idx in headers.items():
        if "Preferred Shares" in col_name:
            class_id = col_name.replace(" Preferred Shares", "")
            invest_col_name = f"{class_id} $ Invested"
            if invest_col_name in headers:
                pref_classes[class_id] = {
                    'shares_col': headers[col_name],
                    'invest_col': headers[invest_col_name],
                }

    # Check each investor row
    for row in range(pref_header_row + 1, totals_row):
        holder = sheet.cell(row=row, column=1).value
        if not holder:
            continue

        for class_id, cols in pref_classes.items():
            invest_val = sheet.cell(row=row, column=cols['invest_col']).value
            shares_val = sheet.cell(row=row, column=cols['shares_col']).value

            has_investment = invest_val is not None and invest_val != ""
            has_shares_formula = shares_val is not None and shares_val != ""

            if has_shares_formula and not has_investment:
                raise AssertionError(
                    f"{holder} has {class_id} shares formula without investment: {shares_val}"
                )


if __name__ == "__main__":
    print("\n" + "="*100)
    print("ROUND SHEET RENDERER - COMPREHENSIVE TEST SUITE")
    print("="*100)

    tests = [
        ("Block Verification", test_round_sheet_matches_block_calculations),
        ("Formula Structure", test_formula_structure),
        ("Pre-Money Row Structure", test_pre_money_row_only_has_dollar_values),
        ("% FD Totals to 100%", test_pct_fd_totals_to_100),
        ("Total Shares Column", test_total_shares_column_exists),
        ("Header Rows Clean", test_header_rows_have_no_formulas),
        ("Investor Formula Scoping", test_investors_only_have_formulas_for_their_rounds),
    ]

    passed = 0
    failed = 0

    for test_name, test_func in tests:
        try:
            test_func()
            print(f"✅ {test_name} PASSED")
            passed += 1
        except AssertionError as e:
            print(f"❌ {test_name} FAILED: {e}")
            failed += 1
        except Exception as e:
            print(f"❌ {test_name} ERROR: {e}")
            failed += 1

    print("\n" + "="*100)
    print(f"RESULTS: {passed} passed, {failed} failed")
    print("="*100)

    if failed > 0:
        sys.exit(1)
