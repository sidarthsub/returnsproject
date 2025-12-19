"""
Excel Formula Validation Helper

Layer 2 testing: Validates that Excel formulas evaluate to expected values
using the 'formulas' library (a Python Excel formula engine with better function support).

Usage:
    validator = ExcelValidator("workbook.xlsx")

    # Check single cell value
    validator.assert_value("Sheet1!B39", expected=254902, tolerance=0.01)

    # Check value is in range
    validator.assert_in_range("Sheet1!B39", min_val=250000, max_val=260000)

    # Get calculated value
    value = validator.get_value("Sheet1!B39")

    # Check formula structure (regex)
    validator.assert_formula_matches("Sheet1!B39", pattern=r"=IFERROR\\(MAX\\(.*\\),0\\)")

Note: Uses 'formulas' library which supports IFERROR, XLOOKUP, IFNA, etc.
      Falls back to xlcalculator if formulas library is not available.
"""

from pathlib import Path
from typing import Union, Optional, Pattern
import re
from decimal import Decimal

from openpyxl import load_workbook

# Try to import formulas library first (better Excel support)
try:
    import formulas
    HAS_FORMULAS = True
    HAS_XLCALCULATOR = False
except ImportError:
    HAS_FORMULAS = False
    # Fall back to xlcalculator
    try:
        from xlcalculator import ModelCompiler, Evaluator, Model
        HAS_XLCALCULATOR = True
    except ImportError:
        HAS_XLCALCULATOR = False


class FormulaEvaluationError(Exception):
    """Raised when formula cannot be evaluated"""
    pass


class ExcelValidator:
    """Helper for validating Excel workbook formulas and values"""

    def __init__(self, workbook_path: Union[str, Path], engine: str = "auto"):
        """
        Initialize validator with workbook path

        Args:
            workbook_path: Path to Excel workbook (.xlsx)
            engine: Formula engine to use: "formulas", "xlcalculator", or "auto" (default)
                   "auto" tries formulas first, falls back to xlcalculator
        """
        if not HAS_FORMULAS and not HAS_XLCALCULATOR:
            raise ImportError(
                "No formula evaluation library available. Install with:\n"
                "  pip install formulas  (recommended - better Excel support)\n"
                "  pip install xlcalculator  (alternative)"
            )

        self.workbook_path = Path(workbook_path)
        if not self.workbook_path.exists():
            raise FileNotFoundError(f"Workbook not found: {self.workbook_path}")

        # Determine engine
        if engine == "auto":
            self.engine = "formulas" if HAS_FORMULAS else "xlcalculator"
        else:
            self.engine = engine
            if engine == "formulas" and not HAS_FORMULAS:
                raise ImportError("formulas library not installed. Install with: pip install formulas")
            if engine == "xlcalculator" and not HAS_XLCALCULATOR:
                raise ImportError("xlcalculator library not installed. Install with: pip install xlcalculator")

        # Load workbook for formula inspection
        self.wb = load_workbook(self.workbook_path, data_only=False)

        # Initialize formula evaluator (lazy-loaded)
        self._excel_model = None
        self._calculated_results = None
        self._xlcalc_evaluator = None
        self._evaluation_errors = {}

    @property
    def excel_model(self):
        """Lazy-load formulas library Excel model"""
        if self._excel_model is None and self.engine == "formulas":
            try:
                self._excel_model = formulas.ExcelModel().loads(str(self.workbook_path)).finish()
            except Exception as e:
                raise FormulaEvaluationError(
                    f"Failed to initialize formulas Excel model: {e}\n"
                    f"This may be due to unsupported Excel features."
                )
        return self._excel_model

    @property
    def calculated_results(self):
        """Lazy-calculate all formulas (formulas library)"""
        if self._calculated_results is None and self.engine == "formulas":
            try:
                self._calculated_results = self.excel_model.calculate()
            except Exception as e:
                raise FormulaEvaluationError(f"Failed to calculate formulas: {e}")
        return self._calculated_results

    @property
    def xlcalc_evaluator(self):
        """Lazy-load xlcalculator evaluator"""
        if self._xlcalc_evaluator is None and self.engine == "xlcalculator":
            try:
                compiler = ModelCompiler()
                model = compiler.read_and_parse_archive(str(self.workbook_path))
                self._xlcalc_evaluator = Evaluator(model)
            except Exception as e:
                raise FormulaEvaluationError(
                    f"Failed to initialize xlcalculator: {e}\n"
                    f"This may be due to unsupported Excel features."
                )
        return self._xlcalc_evaluator

    def get_formula(self, cell_ref: str) -> Optional[str]:
        """
        Get formula from a cell

        Args:
            cell_ref: Cell reference like "Sheet1!B39" or "B39" (uses active sheet)

        Returns:
            Formula string or None if cell has no formula
        """
        sheet_name, cell_addr = self._parse_cell_ref(cell_ref)
        sheet = self.wb[sheet_name]
        return sheet[cell_addr].value if isinstance(sheet[cell_addr].value, str) and sheet[cell_addr].value.startswith('=') else None

    def get_value(self, cell_ref: str, allow_errors: bool = False) -> Union[float, int, str, None]:
        """
        Evaluate formula and get calculated value

        Args:
            cell_ref: Cell reference like "Sheet1!B39"
            allow_errors: If True, return None on error instead of raising

        Returns:
            Calculated value

        Raises:
            FormulaEvaluationError: If formula cannot be evaluated (unless allow_errors=True)
        """
        # Parse cell reference
        sheet_name, cell_addr = self._parse_cell_ref(cell_ref)

        try:
            if self.engine == "formulas":
                # formulas library: calculate all, then get specific cell
                results = self.calculated_results

                # formulas library uses keys like: '[workbook.xlsx]SHEETNAME'!CellAddr
                # Try various key formats
                workbook_name = self.workbook_path.name
                possible_keys = [
                    f"'[{workbook_name}]{sheet_name.upper()}'!{cell_addr}",
                    f"'[{workbook_name}]{sheet_name}'!{cell_addr}",
                    f"'{sheet_name.upper()}'!{cell_addr}",
                    f"'{sheet_name}'!{cell_addr}",
                    f"{sheet_name}!{cell_addr}",
                ]

                result = None
                for key in possible_keys:
                    if key in results:
                        result_obj = results[key]
                        # Extract value from Ranges object
                        if hasattr(result_obj, 'value'):
                            # Ranges object has .value attribute
                            val = result_obj.value
                            # value is often a 2D array [[value]]
                            if isinstance(val, list) and len(val) > 0:
                                if isinstance(val[0], list) and len(val[0]) > 0:
                                    result = val[0][0]
                                else:
                                    result = val[0]
                            else:
                                result = val
                        else:
                            result = result_obj
                        break

                if result is None:
                    # No result found - might be a constant cell
                    # Try reading directly from worksheet
                    cell_value = self.wb[sheet_name][cell_addr].value
                    if cell_value and not str(cell_value).startswith('='):
                        result = cell_value
            else:
                # xlcalculator: evaluate specific cell
                full_ref = f"{sheet_name}!{cell_addr}"
                result = self.xlcalc_evaluator.evaluate(full_ref)

            # Handle numpy types
            if hasattr(result, 'item'):
                result = result.item()

            return result
        except Exception as e:
            error_msg = f"Failed to evaluate {sheet_name}!{cell_addr}: {e}"
            self._evaluation_errors[f"{sheet_name}!{cell_addr}"] = str(e)

            if allow_errors:
                return None
            raise FormulaEvaluationError(error_msg)

    def assert_value(
        self,
        cell_ref: str,
        expected: Union[float, int, Decimal],
        tolerance: float = 0.0,
        rel_tolerance: Optional[float] = None
    ):
        """
        Assert that cell evaluates to expected value

        Args:
            cell_ref: Cell reference like "Sheet1!B39"
            expected: Expected value
            tolerance: Absolute tolerance (default: 0 for exact match)
            rel_tolerance: Relative tolerance as fraction (e.g., 0.01 for ±1%)
                          If specified, overrides tolerance

        Raises:
            AssertionError: If value doesn't match
        """
        actual = self.get_value(cell_ref)

        if actual is None:
            raise AssertionError(f"{cell_ref} evaluated to None (expected {expected})")

        expected = float(expected)
        actual = float(actual)

        # Calculate tolerance
        if rel_tolerance is not None:
            tolerance = abs(expected * rel_tolerance)

        diff = abs(actual - expected)

        if diff > tolerance:
            pct_diff = (diff / expected * 100) if expected != 0 else float('inf')
            raise AssertionError(
                f"{cell_ref}: Expected {expected:,.2f}, got {actual:,.2f}\n"
                f"  Difference: {diff:,.2f} ({pct_diff:.2f}%)\n"
                f"  Tolerance: {tolerance:,.2f}"
            )

    def assert_in_range(
        self,
        cell_ref: str,
        min_val: Union[float, int, Decimal, None] = None,
        max_val: Union[float, int, Decimal, None] = None
    ):
        """
        Assert that cell value is within range

        Args:
            cell_ref: Cell reference
            min_val: Minimum value (inclusive), or None for no minimum
            max_val: Maximum value (inclusive), or None for no maximum

        Raises:
            AssertionError: If value is out of range
        """
        actual = self.get_value(cell_ref)

        if actual is None:
            raise AssertionError(f"{cell_ref} evaluated to None")

        actual = float(actual)

        if min_val is not None and actual < float(min_val):
            raise AssertionError(
                f"{cell_ref}: Expected >= {min_val:,.2f}, got {actual:,.2f}"
            )

        if max_val is not None and actual > float(max_val):
            raise AssertionError(
                f"{cell_ref}: Expected <= {max_val:,.2f}, got {actual:,.2f}"
            )

    def assert_formula_matches(
        self,
        cell_ref: str,
        pattern: Union[str, Pattern],
        exact: bool = False
    ):
        """
        Assert that cell formula matches pattern

        Args:
            cell_ref: Cell reference
            pattern: Regex pattern or exact string to match
            exact: If True, formula must equal pattern exactly (no regex)

        Raises:
            AssertionError: If formula doesn't match
        """
        formula = self.get_formula(cell_ref)

        if formula is None:
            raise AssertionError(f"{cell_ref} has no formula")

        if exact:
            if formula != pattern:
                raise AssertionError(
                    f"{cell_ref} formula mismatch:\n"
                    f"  Expected: {pattern}\n"
                    f"  Actual:   {formula}"
                )
        else:
            if isinstance(pattern, str):
                pattern = re.compile(pattern)

            if not pattern.search(formula):
                raise AssertionError(
                    f"{cell_ref} formula doesn't match pattern:\n"
                    f"  Pattern: {pattern.pattern}\n"
                    f"  Formula: {formula}"
                )

    def assert_no_errors(self, sheet_name: Optional[str] = None):
        """
        Assert that no cells contain Excel errors (#REF!, #VALUE!, etc.)

        Args:
            sheet_name: Sheet to check, or None for all sheets

        Raises:
            AssertionError: If any errors found
        """
        sheets_to_check = [sheet_name] if sheet_name else self.wb.sheetnames
        errors = []

        for sname in sheets_to_check:
            sheet = self.wb[sname]
            for row in sheet.iter_rows():
                for cell in row:
                    if isinstance(cell.value, str) and cell.value.startswith('#'):
                        errors.append(f"{sname}!{cell.coordinate}: {cell.value}")

        if errors:
            raise AssertionError(
                f"Found {len(errors)} error(s) in workbook:\n" +
                "\n".join(f"  - {e}" for e in errors)
            )

    def get_evaluation_errors(self) -> dict:
        """
        Get dictionary of cells that failed to evaluate

        Returns:
            Dict mapping cell reference to error message
        """
        return self._evaluation_errors.copy()

    def _parse_cell_ref(self, cell_ref: str) -> tuple[str, str]:
        """
        Parse cell reference into (sheet_name, cell_address)

        Args:
            cell_ref: "Sheet1!B39" or "B39" (uses active sheet)

        Returns:
            (sheet_name, cell_address) tuple
        """
        if '!' in cell_ref:
            sheet_name, cell_addr = cell_ref.split('!', 1)
            # Remove quotes if present
            sheet_name = sheet_name.strip("'\"")
        else:
            sheet_name = self.wb.sheetnames[0]
            cell_addr = cell_ref

        return sheet_name, cell_addr

    def close(self):
        """Close workbook"""
        self.wb.close()

    def __enter__(self):
        """Context manager support"""
        return self

    def __exit__(self, exc_type, exc_val, exc_tb):
        """Context manager support"""
        self.close()


# Convenience function for quick validation
def validate_formulas(workbook_path: Union[str, Path],
                     validations: dict[str, dict]) -> list[str]:
    """
    Run multiple validations on a workbook

    Args:
        workbook_path: Path to workbook
        validations: Dict mapping cell refs to validation config:
            {
                "Sheet1!B39": {
                    "value": 254902,
                    "tolerance": 0.01,
                },
                "Sheet1!B40": {
                    "min": 1000,
                    "max": 2000,
                },
                "Sheet1!B41": {
                    "formula_pattern": r"=SUM\\(.*\\)"
                }
            }

    Returns:
        List of error messages (empty if all passed)
    """
    errors = []

    with ExcelValidator(workbook_path) as validator:
        for cell_ref, config in validations.items():
            try:
                if "value" in config:
                    validator.assert_value(
                        cell_ref,
                        config["value"],
                        tolerance=config.get("tolerance", 0.0),
                        rel_tolerance=config.get("rel_tolerance")
                    )

                if "min" in config or "max" in config:
                    validator.assert_in_range(
                        cell_ref,
                        min_val=config.get("min"),
                        max_val=config.get("max")
                    )

                if "formula_pattern" in config:
                    validator.assert_formula_matches(
                        cell_ref,
                        config["formula_pattern"]
                    )

            except (AssertionError, FormulaEvaluationError) as e:
                errors.append(str(e))

    return errors
